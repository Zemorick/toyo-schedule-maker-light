#!/usr/bin/env python3
"""Toyo Restaurant Schedule Maker (Light) - Desktop Application

Light edition: identical to the full app but with the OCR-based
"Import from Photo" availability feature removed. No PaddleOCR /
OpenCV / numpy dependencies, no separate OCR venv, no ocr_worker.py.
"""

import json
import os
import sys
import copy
import math
import random
from datetime import datetime, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    messagebox.showerror("Missing Dependency",
                         "openpyxl is required.\nRun setup.sh first.")
    sys.exit(1)

try:
    from fpdf import FPDF
except ImportError:
    FPDF = None  # PDF export disabled but app still works

HAS_OCR = False  # Light edition: photo import removed

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
DAYS = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
SHIFTS = ["morning", "mid", "night"]

MORNING_SIDEWORK = ["A", "B", "C", "D"]
MORNING_EXTRA_ORDER = ["C", "A", "B"]  # doubles in this order
NIGHT_SIDEWORK = ["A", "B", "C", "D", "E", "F", "G", "H"]
NIGHT_EXTRA_ORDER = ["F", "H", "G", "E", "D", "C", "B", "A"]  # doubles in this order

# Hibachi shift targets per week. Most servers should hit 2; senior staff
# (Andy, Dian, Leony) max out at 1 hibachi/week so they stay on regular sidework.
HIBACHI_TARGET_DEFAULT = 2
HIBACHI_TARGET_SENIOR = 1
HIBACHI_SENIOR_NAMES = {"Andy", "Dian", "Leony"}


def hibachi_target(name):
    return HIBACHI_TARGET_SENIOR if name in HIBACHI_SENIOR_NAMES else HIBACHI_TARGET_DEFAULT


def name_first_shift_in_day(day_dict, name):
    """Return 'lunch' if `name` first appears in this day's lunch lists,
    else 'dinner' if they appear in dinner, else None. Used to decide
    where the midshift '*' marker shows — only on the FIRST shift of the
    day so a person who works lunch + dinner gets one marker, not two."""
    if (name in day_dict.get("lunch_servers", [])
            or name in day_dict.get("hibachi_lunch_servers", [])):
        return "lunch"
    if (name in day_dict.get("dinner_servers", [])
            or name in day_dict.get("hibachi_servers", [])):
        return "dinner"
    return None

LUNCH_SERVER_TIME = "(10:15-2:15)"
LUNCH_HOST_TIMES = ["(10:15-4:15)", "(10:15-2:15)"]
LUNCH_MANAGER_TIME = "(10:15-4:15) "
MIDSHIFT_TIME = "2:00-4:30pm"

DINNER_SERVER_TIME = "(4:00 - close)"
DINNER_HOST_TIMES_WEEKDAY = ["(4:15-8:15)", "(4:15-9:15)", "(4:15-9:45)"]
DINNER_HOST_TIMES_FRIDAY = ["(4:15-9:15)", "(4:15-9:45)", "( 4:15-10:15)"]
DINNER_HOST_TIMES_WEEKEND = ["(4:15-9:15)", "(4:15-10:15)", "( 4:15-10:15)"]
DINNER_HOST_TIMES_SUNDAY = ["(2:15 -:8:45)", "(3:15-9:15)", "(4:15-9:45)"]
DINNER_MANAGER_TIME_WEEKDAY = "(4:30-10:00)"
DINNER_MANAGER_TIME_WEEKEND = "(4:15-10:30)"

APP_DIR = Path(__file__).parent
DATA_DIR = APP_DIR / "toyo_data"

# Employee font colors for Excel export (ARGB format)
# Colors extracted from old schedule files — shared colors are intentional
STAFF_COLORS = {
    "Addison": "FFDE00CE",
    "Andy": "FFFF0000",
    "Arabella": "FF0070C0",
    "Ashlyn": "FF00B0F0",
    "Catie Grey": "FF4F81BD",
    "Dian": "FFC00000",
    "Hannah": "FF7B4B23",
    "Jazz": "FF9900FF",
    "Kamryn": "FF748C42",
    "Leony": "FF3421BD",
    "Lilly": "FFF519DB",
    "Maddie": "FF9900FF",
    "Makayla": "FF3B608D",
    "Maria": "FFC00000",
    "Olivia": "FF4D5D2C",
    "Owen": "FF00B050",
    "Sadie": "FFF519DB",
    "Sam": "FF4F81BD",
    "Trish": "FF7B4B23",
}
# Color pool for new employees (legible, not yellow/white/black)
_COLOR_POOL = [
    "FF0070C0", "FFFF0000", "FFC00000", "FF3421BD", "FF9900FF",
    "FF00B050", "FF00B0F0", "FFDE00CE", "FFF519DB", "FF7B4B23",
    "FF4F81BD", "FF3B608D", "FF748C42", "FF4D5D2C", "FF8064A2",
    "FF4BACC6", "FFE36C09", "FF00B0F0", "FF7030A0", "FF002060",
]

# ---------------------------------------------------------------------------
# Default Staff Roster
# ---------------------------------------------------------------------------
DEFAULT_STAFF = [
    # Servers
    {"name": "Dian", "roles": ["server"], "seniority": 10, "fixed_schedule": True,
     "active": True, "flags": [], "role_preference": "server",
     "default_off": ["MON", "TUE"]},
    {"name": "Leony", "roles": ["server"], "seniority": 10, "fixed_schedule": True,
     "active": True, "flags": [], "role_preference": "server",
     "default_off": ["WED", "THU"]},
    {"name": "Will", "roles": ["server"], "seniority": 5, "fixed_schedule": False,
     "active": True, "flags": ["always_hibachi"], "role_preference": "server"},
    {"name": "Andy", "roles": ["server"], "seniority": 6, "fixed_schedule": False,
     "active": True, "flags": ["fill_in"], "role_preference": "server",
     "preferred_off": ["TUE", "WED", "THU"]},
    {"name": "Winnie", "roles": ["server", "host"], "seniority": 3, "fixed_schedule": False,
     "active": True, "flags": ["emergency_only"], "role_preference": None},
    {"name": "Cindy", "roles": ["server"], "seniority": 3, "fixed_schedule": False,
     "active": True, "flags": ["fill_in"], "role_preference": "server"},
    {"name": "Maddie", "roles": ["server"], "seniority": 5, "fixed_schedule": False,
     "active": True, "flags": [], "role_preference": "server"},
    {"name": "Owen", "roles": ["server"], "seniority": 5, "fixed_schedule": False,
     "active": True, "flags": [], "role_preference": "server"},
    {"name": "Sadie", "roles": ["server"], "seniority": 5, "fixed_schedule": False,
     "active": True, "flags": [], "role_preference": "server"},
    {"name": "Arabella", "roles": ["server"], "seniority": 4, "fixed_schedule": False,
     "active": True, "flags": [], "role_preference": "server"},
    {"name": "Kat", "roles": ["server"], "seniority": 5, "fixed_schedule": False,
     "active": True, "flags": [], "role_preference": "server"},
    {"name": "Trish", "roles": ["server"], "seniority": 5, "fixed_schedule": False,
     "active": True, "flags": [], "role_preference": "server"},
    {"name": "Sam", "roles": ["server"], "seniority": 4, "fixed_schedule": False,
     "active": True, "flags": [], "role_preference": "server"},
    {"name": "Garret", "roles": ["server"], "seniority": 4, "fixed_schedule": False,
     "active": True, "flags": [], "role_preference": "server"},
    {"name": "Collin", "roles": ["server"], "seniority": 5, "fixed_schedule": False,
     "active": True, "flags": [], "role_preference": "server"},
    {"name": "Q", "roles": ["server"], "seniority": 4, "fixed_schedule": False,
     "active": True, "flags": [], "role_preference": "server"},
    {"name": "Bryan", "roles": ["server"], "seniority": 4, "fixed_schedule": False,
     "active": True, "flags": [], "role_preference": "server"},
    # Hosts
    {"name": "Maria", "roles": ["host"], "seniority": 5, "fixed_schedule": False,
     "active": True, "flags": ["no_closing"], "role_preference": "host"},
    {"name": "Makayla", "roles": ["host"], "seniority": 5, "fixed_schedule": False,
     "active": True, "flags": [], "role_preference": "host"},
    {"name": "Catie Grey", "roles": ["host"], "seniority": 4, "fixed_schedule": False,
     "active": True, "flags": [], "role_preference": "host"},
    {"name": "Olivia", "roles": ["host"], "seniority": 8, "fixed_schedule": True,
     "active": True, "flags": ["seniority_priority"], "role_preference": "host",
     "default_availability": {
         "MON": "morning", "TUE": "morning", "WED": "morning",
         "THU": "morning", "FRI": "morning",
     },
     "default_off": ["SAT", "SUN"]},
    {"name": "Lilly", "roles": ["host"], "seniority": 4, "fixed_schedule": False,
     "active": True, "flags": [], "role_preference": "host"},
    {"name": "Ashlyn", "roles": ["host"], "seniority": 4, "fixed_schedule": False,
     "active": True, "flags": [], "role_preference": "host"},
    {"name": "Addison", "roles": ["host"], "seniority": 4, "fixed_schedule": False,
     "active": True, "flags": [], "role_preference": "host"},
    {"name": "Hannah", "roles": ["host"], "seniority": 4, "fixed_schedule": False,
     "active": True, "flags": [], "role_preference": "host"},
    # Dual role
    {"name": "Jazz", "roles": ["server", "host"], "seniority": 5, "fixed_schedule": False,
     "active": True, "flags": [], "role_preference": "server"},
    {"name": "Kamryn", "roles": ["server", "host"], "seniority": 5, "fixed_schedule": False,
     "active": True, "flags": [], "role_preference": "host"},
    {"name": "Abigail", "roles": ["server", "host"], "seniority": 5, "fixed_schedule": False,
     "active": True, "flags": [], "role_preference": "server"},
    # Emergency / special
    {"name": "Ross", "roles": ["server", "host"], "seniority": 3, "fixed_schedule": False,
     "active": True, "flags": ["emergency_only"], "role_preference": None},
    {"name": "Shayne", "roles": ["server"], "seniority": 3, "fixed_schedule": False,
     "active": True, "flags": ["emergency_only"], "role_preference": "server"},
    # Managers
    {"name": "Aaron", "roles": ["manager"], "seniority": 10, "fixed_schedule": True,
     "active": True, "flags": [], "role_preference": "manager",
     "default_availability": {
         "MON": "both", "TUE": "off",
         "WED": "night", "THU": "night",
         "FRI": "both", "SAT": "both", "SUN": "both",
     }},
    {"name": "Chan", "roles": ["manager"], "seniority": 10, "fixed_schedule": True,
     "active": True, "flags": [], "role_preference": "manager",
     "default_availability": {
         "MON": "off", "TUE": "both",
         "WED": "morning", "THU": "morning", "FRI": "morning",
         "SAT": "off", "SUN": "off",
     }},
]

DEFAULT_CONFIG = {
    "managers": {"lunch": "Aaron", "dinner": "Aaron"},
    "staffing": {
        day: {
            "lunch_servers": 3 if day in ("MON", "TUE", "WED", "THU") else 5,
            "lunch_hosts": 2,
            "mid_servers": 1 if day in ("MON", "TUE", "WED", "THU") else 2,
            "dinner_servers": 5 if day in ("MON", "TUE", "WED", "THU") else 8,
            "dinner_hosts": 3,
            "hibachi_lunch": 1,
            "hibachi_dinner": 2 if day in ("FRI", "SAT", "SUN") else 1,
        }
        for day in DAYS
    },
    "export_path": "",
    "week_start": "",
}


# ---------------------------------------------------------------------------
# Data Management
# ---------------------------------------------------------------------------
class DataManager:
    def __init__(self):
        DATA_DIR.mkdir(exist_ok=True)
        (DATA_DIR / "schedule_history").mkdir(exist_ok=True)
        self.staff_file = DATA_DIR / "staff.json"
        self.config_file = DATA_DIR / "config.json"
        self.availability_file = DATA_DIR / "availability.json"
        self.staff = self._load_staff()
        self.config = self._load_config()
        self.availability = self._load_availability()

    def _load_staff(self):
        if self.staff_file.exists():
            with open(self.staff_file) as f:
                return json.load(f)
        return copy.deepcopy(DEFAULT_STAFF)

    def _load_config(self):
        if self.config_file.exists():
            with open(self.config_file) as f:
                saved = json.load(f)
            # Merge with defaults for any missing keys
            merged = copy.deepcopy(DEFAULT_CONFIG)
            merged.update(saved)
            for day in DAYS:
                if day in saved.get("staffing", {}):
                    merged["staffing"][day].update(saved["staffing"][day])
            return merged
        return copy.deepcopy(DEFAULT_CONFIG)

    def _load_availability(self):
        if self.availability_file.exists():
            with open(self.availability_file) as f:
                return json.load(f)
        return {}

    def save_staff(self):
        with open(self.staff_file, "w") as f:
            json.dump(self.staff, f, indent=2)

    def save_config(self):
        with open(self.config_file, "w") as f:
            json.dump(self.config, f, indent=2)

    def save_availability(self):
        with open(self.availability_file, "w") as f:
            json.dump(self.availability, f, indent=2)

    def get_active_staff(self, role=None):
        result = [s for s in self.staff if s["active"]]
        if role:
            result = [s for s in result if role in s["roles"]]
        return result

    def get_staff_by_name(self, name):
        for s in self.staff:
            if s["name"] == name:
                return s
        return None


# ---------------------------------------------------------------------------
# Schedule Generator
# ---------------------------------------------------------------------------
class ScheduleGenerator:
    def __init__(self, data: DataManager):
        self.data = data
        self.schedule = {
            day: {
                "lunch_servers": [],
                "lunch_hosts": [],
                "lunch_manager": "",
                "mid_servers": [],
                "dinner_servers": [],
                "dinner_hosts": [],
                "dinner_manager": "",
                "hibachi_lunch_servers": [],
                "hibachi_servers": [],
                "b_side_lunch": [],
                "b_side_dinner": [],
            }
            for day in DAYS
        }
        self.shift_counts = {}  # name -> total shifts
        self.hibachi_counts = {}  # name -> hibachi shifts this week
        self.closing_counts = {}  # name -> closing host shifts this week
        self.warnings = []

    def generate(self):
        self.shift_counts = {}
        self.hibachi_counts = {}
        self.closing_counts = {}
        self.warnings = []
        avail = self.data.availability
        config = self.data.config

        # Reset schedule
        for day in DAYS:
            for key in self.schedule[day]:
                if isinstance(self.schedule[day][key], list):
                    self.schedule[day][key] = []
                else:
                    self.schedule[day][key] = ""

        # Step 1: Place managers based on their availability
        for day in DAYS:
            aaron_avail = self._get_availability("Aaron", day)
            chan_avail = self._get_availability("Chan", day)

            aaron_morning = aaron_avail in ("morning", "both")
            aaron_night = aaron_avail in ("night", "both")
            chan_morning = chan_avail in ("morning", "both")
            chan_night = chan_avail in ("night", "both")

            # Lunch manager(s)
            lunch_mgrs = []
            if aaron_morning:
                lunch_mgrs.append("Aaron")
            if chan_morning:
                lunch_mgrs.append("Chan")
            if not lunch_mgrs:
                lunch_mgrs.append(config["managers"].get("lunch", "Aaron"))
            self.schedule[day]["lunch_manager"] = "/".join(lunch_mgrs)

            # Dinner manager(s)
            dinner_mgrs = []
            if aaron_night:
                dinner_mgrs.append("Aaron")
            if chan_night:
                dinner_mgrs.append("Chan")
            if not dinner_mgrs:
                dinner_mgrs.append(config["managers"].get("dinner", "Aaron"))
            self.schedule[day]["dinner_manager"] = "/".join(dinner_mgrs)

        # Step 2: Place fixed schedule staff (Dian & Leony)
        self._place_fixed_staff(avail, config)

        # Step 3: Place Will (always hibachi)
        self._place_will(avail, config)

        # Step 4: Place hosts with seniority (Olivia first)
        self._place_hosts(avail, config)

        # Step 5: Place remaining servers
        self._place_servers(avail, config)

        # Step 6: Balance hibachi (ensure ~2 per server per week)
        self._balance_hibachi(config)

        # Step 7: Assign sidework letters (after hibachi so hibachi gets last letters)
        self._assign_sidework_letters()

        # Step 8: Check shortages (after all placement is done)
        self._check_shortages(config)

        return self.schedule, self.warnings

    def _get_availability(self, name, day):
        avail = self.data.availability
        if name in avail and day in avail[name]:
            return avail[name][day]
        # Fall back to default_availability for fixed-schedule staff (Aaron, Chan, Dian, Leony)
        staff = self.data.get_staff_by_name(name)
        if staff and staff.get("fixed_schedule"):
            default_avail = staff.get("default_availability", {})
            if day in default_avail:
                return default_avail[day]
            default_off = staff.get("default_off", [])
            return "off" if day in default_off else "both"
        return "off"

    def _can_work(self, name, day, shift):
        a = self._get_availability(name, day)
        if a == "both":
            return True
        if a == "morning" and shift in ("morning", "mid"):
            return True
        if a == "night" and shift in ("night", "mid"):
            return True
        return False

    def _is_assigned(self, name, day, shift=None):
        """Check if staff is already assigned. If shift is given, only check that shift."""
        s = self.schedule[day]
        lunch_keys = ["lunch_servers", "lunch_hosts", "hibachi_lunch_servers"]
        dinner_keys = ["dinner_servers", "dinner_hosts", "hibachi_servers"]
        if shift == "morning":
            keys = lunch_keys
        elif shift == "night":
            keys = dinner_keys
        else:
            keys = lunch_keys + ["mid_servers"] + dinner_keys
        for key in keys:
            if name in s[key]:
                return True
        return False

    def _add_shift(self, name):
        self.shift_counts[name] = self.shift_counts.get(name, 0) + 1

    def _get_shift_count(self, name):
        return self.shift_counts.get(name, 0)

    def _place_fixed_staff(self, avail, config):
        for name in ["Dian", "Leony"]:
            staff = self.data.get_staff_by_name(name)
            if not staff or not staff["active"]:
                continue
            for day in DAYS:
                a = self._get_availability(name, day)
                if a == "off":
                    continue
                staffing = config["staffing"][day]
                # Place in lunch if available for morning and lunch needs servers
                if a in ("morning", "both"):
                    if len(self.schedule[day]["lunch_servers"]) < staffing["lunch_servers"]:
                        self.schedule[day]["lunch_servers"].append(name)
                        self._add_shift(name)
                # Place in dinner if available for night
                if a in ("night", "both"):
                    if len(self.schedule[day]["dinner_servers"]) < staffing["dinner_servers"]:
                        self.schedule[day]["dinner_servers"].append(name)
                        self._add_shift(name)

    def _place_will(self, avail, config):
        staff = self.data.get_staff_by_name("Will")
        if not staff or not staff["active"]:
            return
        for day in DAYS:
            a = self._get_availability("Will", day)
            if a == "off":
                continue
            staffing = config["staffing"][day]
            if a in ("night", "both"):
                self.schedule[day]["hibachi_servers"].append("Will")
                self.hibachi_counts["Will"] = self.hibachi_counts.get("Will", 0) + 1
                self._add_shift("Will")
            elif a == "morning":
                if len(self.schedule[day]["lunch_servers"]) < staffing["lunch_servers"]:
                    self.schedule[day]["lunch_servers"].append("Will")
                    self._add_shift("Will")

    def _place_hosts(self, avail, config):
        hosts = [s for s in self.data.get_active_staff("host")
                 if not s["fixed_schedule"] and "emergency_only" not in s["flags"]]
        # Sort by seniority descending
        hosts.sort(key=lambda s: s["seniority"], reverse=True)

        for day in DAYS:
            staffing = config["staffing"][day]
            # Lunch hosts
            needed = staffing["lunch_hosts"]
            candidates = [h for h in hosts
                          if self._can_work(h["name"], day, "morning")
                          and not self._is_assigned(h["name"], day, "morning")]
            # Sort: fewer shifts first (balance), then seniority for ties
            candidates.sort(key=lambda h: (self._get_shift_count(h["name"]), -h["seniority"]))
            # Olivia always gets position 0 (the 10:15-4:15 slot) if available
            olivia = next((h for h in candidates if h["name"] == "Olivia"), None)
            if olivia:
                candidates.remove(olivia)
                candidates.insert(0, olivia)
            for h in candidates[:needed - len(self.schedule[day]["lunch_hosts"])]:
                self.schedule[day]["lunch_hosts"].append(h["name"])
                self._add_shift(h["name"])

            # Dinner hosts — balance closing (last slot)
            needed = staffing["dinner_hosts"]
            already = len(self.schedule[day]["dinner_hosts"])
            slots_to_fill = needed - already
            if slots_to_fill > 0:
                # Candidates who can close (excludes no_closing flag)
                can_close = [h for h in hosts
                             if self._can_work(h["name"], day, "night")
                             and not self._is_assigned(h["name"], day, "night")
                             and "no_closing" not in (self.data.get_staff_by_name(h["name"]) or {}).get("flags", [])]
                # Candidates who can't close (Maria etc.) — only for non-closing slots
                no_close = [h for h in hosts
                            if self._can_work(h["name"], day, "night")
                            and not self._is_assigned(h["name"], day, "night")
                            and "no_closing" in (self.data.get_staff_by_name(h["name"]) or {}).get("flags", [])]

                assigned_names = set()

                # STEP 1: Decide WHO closes — fewest closing_counts first
                # They will be appended LAST to dinner_hosts so they occupy the latest time slot
                closer = None
                if slots_to_fill >= 1 and can_close:
                    closing_sorted = sorted(can_close, key=lambda h: (
                        self.closing_counts.get(h["name"], 0),
                        self._get_shift_count(h["name"]),
                        -h["seniority"]))
                    closer = closing_sorted[0]
                    assigned_names.add(closer["name"])

                # STEP 2: Fill early positions (hostess 1, 2, ...) with non-closers
                # Prefer no_close staff (Maria etc.) so they never end up in the last slot
                non_closing_slots = slots_to_fill - (1 if closer else 0)
                no_close.sort(key=lambda h: (self._get_shift_count(h["name"]), -h["seniority"]))
                can_close.sort(key=lambda h: (self._get_shift_count(h["name"]), -h["seniority"]))
                non_closing_pool = no_close + [h for h in can_close if h["name"] not in assigned_names]
                for h in non_closing_pool:
                    if non_closing_slots <= 0:
                        break
                    if h["name"] not in assigned_names:
                        self.schedule[day]["dinner_hosts"].append(h["name"])
                        self._add_shift(h["name"])
                        assigned_names.add(h["name"])
                        non_closing_slots -= 1

                # STEP 3: Append the closer LAST (takes the 4:15-9:45 slot)
                if closer:
                    self.schedule[day]["dinner_hosts"].append(closer["name"])
                    self._add_shift(closer["name"])
                    self.closing_counts[closer["name"]] = self.closing_counts.get(closer["name"], 0) + 1

        # Check if we need dual-role staff to fill host gaps
        for day in DAYS:
            staffing = config["staffing"][day]
            for shift_key, shift_type, needed_key in [
                ("lunch_hosts", "morning", "lunch_hosts"),
                ("dinner_hosts", "night", "dinner_hosts"),
            ]:
                needed = staffing[needed_key]
                current = len(self.schedule[day][shift_key])
                if current < needed:
                    # Try dual-role staff who prefer hosting
                    dual = [s for s in self.data.get_active_staff()
                            if "host" in s["roles"] and "server" in s["roles"]
                            and s["role_preference"] == "host"
                            and not s["fixed_schedule"]
                            and "emergency_only" not in s["flags"]
                            and self._can_work(s["name"], day, shift_type)
                            and not self._is_assigned(s["name"], day, shift_type)]
                    dual.sort(key=lambda s: self._get_shift_count(s["name"]))
                    for s in dual[:needed - current]:
                        self.schedule[day][shift_key].append(s["name"])
                        self._add_shift(s["name"])

    def _place_servers(self, avail, config):
        all_servers = [s for s in self.data.get_active_staff("server")
                       if not s["fixed_schedule"]
                       and "always_hibachi" not in s["flags"]
                       and "fill_in" not in s["flags"]
                       and "emergency_only" not in s["flags"]
                       and s.get("role_preference") != "host"]

        fill_ins = [s for s in self.data.get_active_staff("server")
                    if "fill_in" in s["flags"]]

        for day in DAYS:
            # Sort fill-ins: deprioritize those who prefer this day off
            def fill_in_sort_key(s):
                preferred_off = s.get("preferred_off", [])
                penalty = 10 if day in preferred_off else 0
                return (penalty, self._get_shift_count(s["name"]))

            staffing = config["staffing"][day]

            # Lunch servers (hibachi is picked from this pool, not additional)
            needed = staffing["lunch_servers"]
            current = len(self.schedule[day]["lunch_servers"])
            if current < needed:
                candidates = [s for s in all_servers
                              if self._can_work(s["name"], day, "morning")
                              and not self._is_assigned(s["name"], day, "morning")]
                # Deprioritize staff who prefer this day off
                def server_sort_key(s):
                    preferred_off = s.get("preferred_off", [])
                    off_penalty = 5 if day in preferred_off else 0
                    return (off_penalty, self._get_shift_count(s["name"]), -s["seniority"])
                candidates.sort(key=server_sort_key)
                for s in candidates[:needed - current]:
                    self.schedule[day]["lunch_servers"].append(s["name"])
                    self._add_shift(s["name"])

            # Still short? Try dual-role staff
            current = len(self.schedule[day]["lunch_servers"])
            if current < needed:
                dual = [s for s in self.data.get_active_staff()
                        if "server" in s["roles"] and s["role_preference"] == "server"
                        and not s["fixed_schedule"]
                        and "fill_in" not in s["flags"]
                        and "emergency_only" not in s["flags"]
                        and self._can_work(s["name"], day, "morning")
                        and not self._is_assigned(s["name"], day, "morning")]
                dual.sort(key=lambda s: self._get_shift_count(s["name"]))
                for s in dual[:needed - current]:
                    self.schedule[day]["lunch_servers"].append(s["name"])
                    self._add_shift(s["name"])

            # Still short? Try fill-ins
            current = len(self.schedule[day]["lunch_servers"])
            if current < needed:
                fi = [s for s in fill_ins
                      if self._can_work(s["name"], day, "morning")
                      and not self._is_assigned(s["name"], day, "morning")]
                fi.sort(key=fill_in_sort_key)
                for s in fi[:needed - current]:
                    self.schedule[day]["lunch_servers"].append(s["name"])
                    self._add_shift(s["name"])

            # Dinner servers (hibachi is picked from this pool, not additional)
            needed = staffing["dinner_servers"]
            current = len(self.schedule[day]["dinner_servers"])
            if current < needed:
                candidates = [s for s in all_servers
                              if self._can_work(s["name"], day, "night")
                              and not self._is_assigned(s["name"], day, "night")]
                def dinner_sort_key(s):
                    preferred_off = s.get("preferred_off", [])
                    off_penalty = 5 if day in preferred_off else 0
                    return (off_penalty, self._get_shift_count(s["name"]), -s["seniority"])
                candidates.sort(key=dinner_sort_key)
                for s in candidates[:needed - current]:
                    self.schedule[day]["dinner_servers"].append(s["name"])
                    self._add_shift(s["name"])

            # Dinner - try dual-role
            current = len(self.schedule[day]["dinner_servers"])
            if current < needed:
                dual = [s for s in self.data.get_active_staff()
                        if "server" in s["roles"] and s["role_preference"] == "server"
                        and not s["fixed_schedule"]
                        and "fill_in" not in s["flags"]
                        and "emergency_only" not in s["flags"]
                        and self._can_work(s["name"], day, "night")
                        and not self._is_assigned(s["name"], day, "night")]
                dual.sort(key=lambda s: self._get_shift_count(s["name"]))
                for s in dual[:needed - current]:
                    self.schedule[day]["dinner_servers"].append(s["name"])
                    self._add_shift(s["name"])

            # Dinner - try fill-ins
            current = len(self.schedule[day]["dinner_servers"])
            if current < needed:
                fi = [s for s in fill_ins
                      if self._can_work(s["name"], day, "night")
                      and not self._is_assigned(s["name"], day, "night")]
                fi.sort(key=fill_in_sort_key)
                for s in fi[:needed - current]:
                    self.schedule[day]["dinner_servers"].append(s["name"])
                    self._add_shift(s["name"])

            # Midshift
            needed_mid = staffing["mid_servers"]
            # Mid servers are selected by manager later, but we can suggest
            # For now leave empty - manager picks in the GUI


    def _check_shortages(self, config):
        """Check for staffing shortages after all placement is done."""
        for day in DAYS:
            staffing = config["staffing"][day]
            for key, label in [("lunch_servers", "Lunch Servers"),
                               ("dinner_servers", "Dinner Servers"),
                               ("lunch_hosts", "Lunch Hosts"),
                               ("dinner_hosts", "Dinner Hosts")]:
                if key == "lunch_hosts":
                    n = staffing["lunch_hosts"]
                elif key == "dinner_hosts":
                    n = staffing["dinner_hosts"]
                elif key == "lunch_servers":
                    n = staffing["lunch_servers"]
                else:
                    n = staffing["dinner_servers"]
                actual = len(self.schedule[day][key])
                # Include hibachi servers in the count
                if key == "dinner_servers":
                    actual += len(self.schedule[day]["hibachi_servers"])
                elif key == "lunch_servers":
                    actual += len(self.schedule[day]["hibachi_lunch_servers"])
                if actual < n:
                    self.warnings.append(
                        f"{day}: Need {n} {label} but only found {actual}")

    def _find_a_holder(self, candidates, priority_chain):
        """Find who should get letter A from a priority chain.
        Returns the first person in priority_chain who is in candidates,
        or None if nobody from the chain is available."""
        for name in priority_chain:
            if name in candidates:
                return name
        return None

    def _assign_sidework_letters(self):
        weekdays = ["MON", "TUE", "WED", "THU"]

        # Morning letter-A priority:
        #   Fri/Sat: Andy -> Dian/Leony -> Abigail -> Bryan -> random
        #   Weekdays: Dian -> Leony -> Abigail -> Bryan -> random
        #   Sun: Dian -> Leony -> Abigail -> Bryan -> random
        MORNING_A_PRIORITY_FRISATAM = ["Andy", "Dian", "Leony", "Abigail", "Bryan"]
        MORNING_A_PRIORITY_DEFAULT = ["Dian", "Leony", "Abigail", "Bryan"]

        # Night letter-A priority:
        #   Andy -> Dian -> Leony -> Sadie -> Bryan -> random
        NIGHT_A_PRIORITY = ["Andy", "Dian", "Leony", "Sadie", "Bryan"]

        for day in DAYS:
            # --- Morning servers get A-D ---
            # B-side staff are excluded from the A-side letter pool entirely so
            # the remaining A,B,C,D... letters compact without gaps. They get
            # their own W/X/Y assignment further down.
            b_lunch_set = set(self.schedule[day].get("b_side_lunch", []))
            servers = [n for n in self.schedule[day]["lunch_servers"] if n not in b_lunch_set]
            hibachi_lunch = list(self.schedule[day]["hibachi_lunch_servers"])

            self.schedule[day]["lunch_server_letters"] = {}

            # Determine morning A holder (from regular servers only)
            if day in ("FRI", "SAT"):
                morning_a = self._find_a_holder(servers, MORNING_A_PRIORITY_FRISATAM)
            else:
                morning_a = self._find_a_holder(servers, MORNING_A_PRIORITY_DEFAULT)

            if morning_a:
                ordered_regular = [morning_a] + [n for n in servers if n != morning_a]
            else:
                ordered_regular = servers

            # Regular servers get letters A, B, C, D... in order
            reg_letters = list(MORNING_SIDEWORK)
            if len(ordered_regular) > len(reg_letters):
                for extra in MORNING_EXTRA_ORDER:
                    reg_letters.append(extra)
                    if len(reg_letters) >= len(ordered_regular):
                        break
            for i, name in enumerate(ordered_regular):
                self.schedule[day]["lunch_server_letters"][name] = reg_letters[i] if i < len(reg_letters) else reg_letters[-1]

            # Hibachi get letters from the hibachi pool (cycles for extras, never A)
            hibachi_pool = ["D", "C"]  # morning hibachi letters — cycle through these
            for i, name in enumerate(hibachi_lunch):
                self.schedule[day]["lunch_server_letters"][name] = hibachi_pool[i % len(hibachi_pool)]

            # B-side (sushi side) overrides letters with W, X, Y in insertion order
            b_side_pool = ["W", "X", "Y"]
            for i, name in enumerate(self.schedule[day].get("b_side_lunch", [])):
                self.schedule[day]["lunch_server_letters"][name] = b_side_pool[i] if i < len(b_side_pool) else b_side_pool[-1]

            # --- Night servers get A-H ---
            # B-side staff excluded from the A-side pool so A,B,C... compacts without gaps.
            b_dinner_set = set(self.schedule[day].get("b_side_dinner", []))
            dinner_regular = [n for n in self.schedule[day]["dinner_servers"] if n not in b_dinner_set]
            dinner_hibachi = list(self.schedule[day]["hibachi_servers"])

            self.schedule[day]["dinner_server_letters"] = {}

            night_a = self._find_a_holder(dinner_regular, NIGHT_A_PRIORITY)

            if night_a:
                ordered_night = [night_a] + [n for n in dinner_regular if n != night_a]
            else:
                ordered_night = dinner_regular

            # Regular servers get letters A, B, C, D, E, F, G, H... in order
            reg_letters = list(NIGHT_SIDEWORK)
            if len(ordered_night) > len(reg_letters):
                for extra in NIGHT_EXTRA_ORDER:
                    reg_letters.append(extra)
                    if len(reg_letters) >= len(ordered_night):
                        break
            for i, name in enumerate(ordered_night):
                self.schedule[day]["dinner_server_letters"][name] = reg_letters[i] if i < len(reg_letters) else reg_letters[-1]

            # Hibachi get letters from the hibachi pool (cycles for extras, never A or B)
            hibachi_pool = ["F", "H", "G", "E"]  # dinner hibachi letters — cycle through
            for i, name in enumerate(dinner_hibachi):
                self.schedule[day]["dinner_server_letters"][name] = hibachi_pool[i % len(hibachi_pool)]

            # B-side (sushi side) overrides letters with W, X, Y in insertion order
            b_side_pool = ["W", "X", "Y"]
            for i, name in enumerate(self.schedule[day].get("b_side_dinner", [])):
                self.schedule[day]["dinner_server_letters"][name] = b_side_pool[i] if i < len(b_side_pool) else b_side_pool[-1]


    def _balance_hibachi(self, config):
        # Each server should get ~2 hibachi shifts per week
        # Will is excluded (always hibachi)
        all_servers = [s["name"] for s in self.data.get_active_staff("server")
                       if "always_hibachi" not in s["flags"]
                       and "emergency_only" not in s["flags"]
                       and s["active"]]

        # Track who has hibachi assignments
        for name in all_servers:
            self.hibachi_counts[name] = 0

        # Assign lunch hibachi — pick from lunch_servers pool
        for day in DAYS:
            staffing = config["staffing"][day]
            needed = staffing.get("hibachi_lunch", 0)
            already = len(self.schedule[day]["hibachi_lunch_servers"])
            needed = needed - already
            if needed <= 0:
                continue
            lunch = self.schedule[day]["lunch_servers"][:]
            lunch_eligible = [n for n in lunch if n in all_servers
                              and self.hibachi_counts.get(n, 0) < 2]
            # Sort: fewest hibachi first, then lowest seniority first (so senior staff stay on regular sidework)
            lunch_eligible.sort(key=lambda n: (
                self.hibachi_counts.get(n, 0),
                -(self.data.get_staff_by_name(n) or {}).get("seniority", 0)))

            picked = lunch_eligible[:needed]
            # Fallback: if not enough with < 2 count, allow anyone (but never exceed 2 for Dian/Leony)
            if len(picked) < needed:
                remaining = needed - len(picked)
                picked_names = set(n for n in picked)
                fallback = [n for n in lunch if n in all_servers and n not in picked_names
                            and (n not in ("Dian", "Leony") or self.hibachi_counts.get(n, 0) < 2)]
                fallback.sort(key=lambda n: self.hibachi_counts.get(n, 0))
                picked.extend(fallback[:remaining])
            for name in picked:
                self.hibachi_counts[name] = self.hibachi_counts.get(name, 0) + 1
                if name in self.schedule[day]["lunch_servers"]:
                    self.schedule[day]["lunch_servers"].remove(name)
                self.schedule[day]["hibachi_lunch_servers"].append(name)

        # Assign dinner hibachi — pick from dinner_servers pool
        for day in DAYS:
            staffing = config["staffing"][day]
            needed = staffing.get("hibachi_dinner", 0)
            already = len(self.schedule[day]["hibachi_servers"])
            needed = needed - already
            if needed <= 0:
                continue
            dinner = self.schedule[day]["dinner_servers"][:]
            dinner_eligible = [n for n in dinner if n in all_servers
                               and self.hibachi_counts.get(n, 0) < 2]
            # Sort: fewest hibachi first, then lowest seniority first
            dinner_eligible.sort(key=lambda n: (
                self.hibachi_counts.get(n, 0),
                -(self.data.get_staff_by_name(n) or {}).get("seniority", 0)))

            picked = dinner_eligible[:needed]
            # Fallback: if not enough with < 2 count, allow anyone (but never exceed 2 for Dian/Leony)
            if len(picked) < needed:
                remaining = needed - len(picked)
                picked_names = set(n for n in picked)
                fallback = [n for n in dinner if n in all_servers and n not in picked_names
                            and (n not in ("Dian", "Leony") or self.hibachi_counts.get(n, 0) < 2)]
                fallback.sort(key=lambda n: self.hibachi_counts.get(n, 0))
                picked.extend(fallback[:remaining])
            for name in picked:
                self.hibachi_counts[name] = self.hibachi_counts.get(name, 0) + 1
                if name in self.schedule[day]["dinner_servers"]:
                    self.schedule[day]["dinner_servers"].remove(name)
                self.schedule[day]["hibachi_servers"].append(name)

        # Second pass: for servers who still have 0 hibachi, try to swap them
        # into a hibachi slot by replacing someone who already has 2
        for name in all_servers:
            if self.hibachi_counts.get(name, 0) >= 1:
                continue
            # Find a day this server works dinner where we can swap
            for day in DAYS:
                if name not in self.schedule[day]["dinner_servers"]:
                    continue
                max_hib = config["staffing"][day].get("hibachi_dinner", 0)
                if len(self.schedule[day]["hibachi_servers"]) > max_hib:
                    continue  # Already over limit, don't swap here
                # Find someone in hibachi on this day with count >= 2
                for hib_name in list(self.schedule[day]["hibachi_servers"]):
                    if hib_name == "Will":
                        continue
                    if self.hibachi_counts.get(hib_name, 0) >= 2:
                        # Swap: move hib_name back to dinner, move name to hibachi
                        self.schedule[day]["hibachi_servers"].remove(hib_name)
                        self.schedule[day]["dinner_servers"].append(hib_name)
                        self.schedule[day]["dinner_servers"].remove(name)
                        self.schedule[day]["hibachi_servers"].append(name)
                        self.hibachi_counts[hib_name] -= 1
                        self.hibachi_counts[name] = self.hibachi_counts.get(name, 0) + 1
                        break
                if self.hibachi_counts.get(name, 0) >= 1:
                    break

        # Check balance and warn
        for name in all_servers:
            count = self.hibachi_counts.get(name, 0)
            if count == 0:
                self.warnings.append(f"{name} has 0 hibachi shifts (target: 2)")
            elif count > 2:
                self.warnings.append(f"{name} has {count} hibachi shifts (target: 2)")


# ---------------------------------------------------------------------------
# Excel Exporter
# ---------------------------------------------------------------------------
class ExcelExporter:
    def __init__(self, schedule, config, week_start_date=None):
        self.schedule = schedule
        self.config = config
        self.week_start = week_start_date

    def export(self, filepath):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Chart2"

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_font = Font(bold=True, size=11)
        day_header_font = Font(bold=True, size=14)
        name_font = Font(size=11)
        time_font = Font(size=11)
        mid_font = Font(size=11, color="0000FF")
        hibachi_fill = PatternFill(start_color="FFFFFF99", end_color="FFFFFF99", fill_type="solid")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"))
        section_top = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="medium"))
        section_bottom = Border(left=Side(style="thin"), right=Side(style="thin"), bottom=Side(style="medium"))
        day_top = Border(right=Side(style="thin"), top=Side(style="medium"))

        def staff_font(name, size=11):
            """Get a Font with the employee's assigned color."""
            color = STAFF_COLORS.get(name)
            if not color:
                # Assign from pool based on hash for consistency
                used = set(STAFF_COLORS.values())
                available = [c for c in _COLOR_POOL if c not in used]
                if not available:
                    available = list(_COLOR_POOL)
                color = available[hash(name) % len(available)]
                STAFF_COLORS[name] = color
            return Font(size=size, color=color[2:])  # strip FF prefix

        # Column widths (match old format)
        ws.column_dimensions["A"].width = 13.4
        for col in ["B", "C", "D", "E", "F", "G"]:
            ws.column_dimensions[col].width = 13.0
        ws.column_dimensions["H"].width = 14.9

        # Row 1: Dates (as datetime with short date format)
        if self.week_start:
            for i, col in enumerate(["B", "C", "D", "E", "F", "G", "H"]):
                d = self.week_start + timedelta(days=i)
                ws[f"{col}1"] = d
                ws[f"{col}1"].font = name_font
                ws[f"{col}1"].number_format = "MM-DD"
                ws[f"{col}1"].alignment = Alignment(horizontal="center")

        # Row 2: Day names
        for i, (col, day) in enumerate(zip(["B", "C", "D", "E", "F", "G", "H"], DAYS)):
            ws[f"{col}2"] = day
            ws[f"{col}2"].font = day_header_font
            ws[f"{col}2"].alignment = Alignment(horizontal="center")

        # LUNCH SERVERS (rows 3-8)
        ws["A4"] = "Servers"
        ws["A4"].font = header_font
        ws["A6"] = LUNCH_SERVER_TIME
        ws["A6"].font = time_font

        for di, day in enumerate(DAYS):
            col = chr(66 + di)  # B-H
            hibachi_lunch = self.schedule[day]["hibachi_lunch_servers"]
            servers = self.schedule[day]["lunch_servers"]
            letters = self.schedule[day].get("lunch_server_letters", {})
            mids = self.schedule[day].get("mid_servers", [])
            b_set = set(self.schedule[day].get("b_side_lunch", []))

            # Sort: B-side on top, then regular A,B,C..., then hibachi (within letter)
            hib_set = set(hibachi_lunch)
            combined = list(servers) + list(hibachi_lunch)
            combined.sort(key=lambda n: (
                0 if n in b_set else 1,
                letters.get(n, "Z"),
                1 if n in hib_set else 0,
            ))

            row = 3
            for name in combined:
                letter = letters.get(name, chr(65 + row - 3))
                # Excel: * marker on every appearance, no color
                mid_mark = "*" if name in mids else ""
                if name in b_set:
                    ws[f"{col}{row}"] = f"<{name} {letter}>{mid_mark}"
                else:
                    ws[f"{col}{row}"] = f"{name} {letter}{mid_mark}"
                ws[f"{col}{row}"].font = staff_font(name)
                ws[f"{col}{row}"].alignment = Alignment(horizontal="center")
                if name in hib_set:
                    ws[f"{col}{row}"].fill = hibachi_fill
                row += 1

        # LUNCH MANAGER (rows 13-14)
        ws["A13"] = "Lunch"
        ws["A13"].font = header_font
        ws["A14"] = "Manager"
        ws["A14"].font = header_font
        for di, day in enumerate(DAYS):
            col = chr(66 + di)
            ws[f"{col}13"] = self.schedule[day]["lunch_manager"]
            ws[f"{col}13"].font = name_font
            ws[f"{col}13"].alignment = Alignment(horizontal="center")
            ws[f"{col}14"] = LUNCH_MANAGER_TIME
            ws[f"{col}14"].font = time_font
            ws[f"{col}14"].alignment = Alignment(horizontal="center")

        # LUNCH HOSTS (rows 15-19)
        ws["A15"] = "Hostess"
        ws["A15"].font = header_font
        ws["A17"] = "Hostess"
        ws["A17"].font = header_font
        ws["A19"] = "Hostess"
        ws["A19"].font = header_font

        for di, day in enumerate(DAYS):
            col = chr(66 + di)
            hosts = self.schedule[day]["lunch_hosts"]
            for hi, name in enumerate(hosts):
                if hi == 0:
                    ws[f"{col}15"] = name
                    ws[f"{col}15"].font = staff_font(name)
                    ws[f"{col}16"] = LUNCH_HOST_TIMES[0]
                    ws[f"{col}16"].font = time_font
                elif hi == 1:
                    ws[f"{col}17"] = name
                    ws[f"{col}17"].font = staff_font(name)
                    ws[f"{col}18"] = LUNCH_HOST_TIMES[1]
                    ws[f"{col}18"].font = time_font
                for r in [15, 16, 17, 18]:
                    if ws[f"{col}{r}"].value:
                        ws[f"{col}{r}"].alignment = Alignment(horizontal="center")

        # Transition time row 20
        time_row_20 = {
            "MON": "(10:15-2:15)", "TUE": "(10:15-4:15)",
            "WED": "(4:00-8:00)", "THU": "(10:15-2:15)",
            "FRI": "(10:15-3:15)", "SAT": "(10:15-2:15)",
            "SUN": "(11:30-3:15)",
        }
        for di, day in enumerate(DAYS):
            col = chr(66 + di)
            ws[f"{col}20"] = time_row_20.get(day, "")
            ws[f"{col}20"].font = time_font
            ws[f"{col}20"].alignment = Alignment(horizontal="center")

        # DINNER SERVERS (rows 21-29)
        ws["A22"] = "Servers"
        ws["A22"].font = header_font
        ws["A23"] = "(s)"
        ws["A23"].font = time_font
        ws["A27"] = DINNER_SERVER_TIME
        ws["A27"].font = time_font

        for di, day in enumerate(DAYS):
            col = chr(66 + di)
            servers = self.schedule[day]["dinner_servers"]
            hibachi = self.schedule[day]["hibachi_servers"]
            letters = self.schedule[day].get("dinner_server_letters", {})
            mids = self.schedule[day].get("mid_servers", [])
            b_set = set(self.schedule[day].get("b_side_dinner", []))

            # Sort: B-side on top, then regular A,B,C..., then hibachi (within letter)
            hib_set = set(hibachi)
            combined = list(servers) + list(hibachi)
            combined.sort(key=lambda n: (
                0 if n in b_set else 1,
                letters.get(n, "Z"),
                1 if n in hib_set else 0,
            ))

            row = 21
            for name in combined:
                letter = letters.get(name, chr(65 + row - 21))
                # Excel: * marker on every appearance, no color
                mid_mark = "*" if name in mids else ""
                if name in b_set:
                    ws[f"{col}{row}"] = f"<{name} {letter}>{mid_mark}"
                else:
                    ws[f"{col}{row}"] = f"{name} {letter}{mid_mark}"
                ws[f"{col}{row}"].font = staff_font(name)
                ws[f"{col}{row}"].alignment = Alignment(horizontal="center")
                if name in hib_set:
                    ws[f"{col}{row}"].fill = hibachi_fill
                row += 1

        # DINNER HOSTS (rows 33-38)
        ws["A33"] = "Hostesses"
        ws["A33"].font = header_font
        ws["A34"] = "(H)"
        ws["A34"].font = time_font

        for di, day in enumerate(DAYS):
            col = chr(66 + di)
            hosts = self.schedule[day]["dinner_hosts"]
            day_idx = di  # 0=MON ... 6=SUN

            if day_idx >= 5:  # SAT, SUN
                times = DINNER_HOST_TIMES_WEEKEND if day != "SUN" else DINNER_HOST_TIMES_SUNDAY
            elif day_idx == 4:  # FRI
                times = DINNER_HOST_TIMES_FRIDAY
            else:
                times = DINNER_HOST_TIMES_WEEKDAY

            for hi, name in enumerate(hosts):
                name_row = 33 + hi * 2
                time_row = 34 + hi * 2
                ws[f"{col}{name_row}"] = name
                ws[f"{col}{name_row}"].font = staff_font(name)
                ws[f"{col}{name_row}"].alignment = Alignment(horizontal="center")
                if hi < len(times):
                    ws[f"{col}{time_row}"] = times[hi]
                    ws[f"{col}{time_row}"].font = time_font
                    ws[f"{col}{time_row}"].alignment = Alignment(horizontal="center")

        # DINNER MANAGER (rows 39-40)
        ws["A39"] = "Dinner"
        ws["A39"].font = header_font
        ws["A40"] = "Manager"
        ws["A40"].font = header_font
        for di, day in enumerate(DAYS):
            col = chr(66 + di)
            ws[f"{col}39"] = self.schedule[day]["dinner_manager"]
            ws[f"{col}39"].font = name_font
            ws[f"{col}39"].alignment = Alignment(horizontal="center")
            t = DINNER_MANAGER_TIME_WEEKEND if di >= 4 else DINNER_MANAGER_TIME_WEEKDAY
            ws[f"{col}40"] = t
            ws[f"{col}40"].font = time_font
            ws[f"{col}40"].alignment = Alignment(horizontal="center")

        # Legend
        ws["A41"].fill = hibachi_fill
        ws["B41"] = "Hibachi Servers"
        ws["B41"].font = Font(size=11)
        ws["A42"] = "*"
        ws["A42"].font = Font(bold=True, size=11)
        ws["B42"] = f"Mid shift server(s)  {MIDSHIFT_TIME}"
        ws["B42"].font = Font(size=11)
        ws["A43"].fill = PatternFill(start_color="FFE5B8B7", end_color="FFE5B8B7", fill_type="solid")
        ws["B43"] = "Training shift"
        ws["B43"].font = Font(size=11)

        # --- Apply borders and extend hibachi yellow fill ---
        for col in ["B", "C", "D", "E", "F", "G", "H"]:
            # Row 2: day header top border
            ws[f"{col}2"].border = day_top

            # Lunch servers (rows 3-12): thin borders, section top on row 3
            ws[f"{col}3"].border = section_top
            for r in range(4, 13):
                ws[f"{col}{r}"].border = thin_border

            # Extend lunch hibachi yellow fill down to row 12
            # Find the last hibachi row in this column
            found_hibachi = False
            for r in range(3, 13):
                cell = ws[f"{col}{r}"]
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb == "FFFFFF99":
                    found_hibachi = True
                elif found_hibachi:
                    # Fill empty rows below hibachi with yellow
                    if not cell.value:
                        cell.fill = hibachi_fill
                    elif cell.fill != hibachi_fill:
                        # Next data cell without hibachi = stop
                        break

            # Lunch manager (row 13-14): section top on 13
            ws[f"{col}13"].border = section_top
            ws[f"{col}14"].border = thin_border

            # Lunch hosts (rows 15-19): section top on 15, 17, 19
            ws[f"{col}15"].border = section_top
            ws[f"{col}16"].border = thin_border
            ws[f"{col}17"].border = section_top
            ws[f"{col}18"].border = thin_border
            ws[f"{col}19"].border = thin_border

            # Transition row 20: bottom border
            ws[f"{col}20"].border = section_bottom

            # Dinner servers (rows 21-32): section top on 21
            ws[f"{col}21"].border = section_top
            for r in range(22, 33):
                ws[f"{col}{r}"].border = thin_border

            # Extend dinner hibachi yellow fill down to row 32
            found_hibachi = False
            for r in range(21, 33):
                cell = ws[f"{col}{r}"]
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb == "FFFFFF99":
                    found_hibachi = True
                elif found_hibachi:
                    if not cell.value:
                        cell.fill = hibachi_fill
                    elif cell.fill != hibachi_fill:
                        break

            # Dinner hosts (rows 33-38): section top on 33, 35, 37
            ws[f"{col}33"].border = section_top
            ws[f"{col}34"].border = thin_border
            ws[f"{col}35"].border = section_top
            ws[f"{col}36"].border = thin_border
            ws[f"{col}37"].border = section_top
            ws[f"{col}38"].border = thin_border

            # Dinner manager (rows 39-40): section top on 39, section bottom on 40
            ws[f"{col}39"].border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="medium"), bottom=Side(style="medium"))
            ws[f"{col}40"].border = section_bottom

        wb.save(filepath)
        return filepath


# ---------------------------------------------------------------------------
# PDF Exporter
# ---------------------------------------------------------------------------
class PDFExporter:
    def __init__(self, schedule, config, week_start_date=None):
        self.schedule = schedule
        self.config = config
        self.week_start = week_start_date

    def export(self, filepath):
        if FPDF is None:
            raise RuntimeError("fpdf2 not installed. Run setup.sh.")

        pdf = FPDF(orientation="L", unit="mm", format="A4")
        pdf.add_page()
        pdf.set_auto_page_break(auto=False)

        col_w = [30] + [35] * 7  # A column + 7 day columns
        x_start = 5
        y_start = 10

        def draw_row(y, texts, bold=False, size=7):
            pdf.set_font("Helvetica", "B" if bold else "", size)
            for i, txt in enumerate(texts):
                x = x_start + sum(col_w[:i])
                pdf.set_xy(x, y)
                pdf.cell(col_w[i], 5, str(txt), border=1, align="C")
            return y + 5

        # Title
        title = "Toyo Schedule"
        if self.week_start:
            end = self.week_start + timedelta(days=6)
            title += f"  {self.week_start.strftime('%b %d')} - {end.strftime('%b %d, %Y')}"
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_xy(x_start, y_start)
        pdf.cell(sum(col_w), 7, title, align="C")
        y = y_start + 9

        # Day headers
        headers = [""] + DAYS
        if self.week_start:
            headers = [""] + [(self.week_start + timedelta(days=i)).strftime("%a %m/%d") for i in range(7)]
        y = draw_row(y, headers, bold=True, size=8)

        # LUNCH SECTION
        y = draw_row(y, ["LUNCH SERVERS", "", "", "", "", "", "", ""], bold=True, size=7)

        # Find max lunch servers across days (including hibachi lunch)
        max_lunch = max(len(self.schedule[d]["lunch_servers"]) + len(self.schedule[d]["hibachi_lunch_servers"]) for d in DAYS) if DAYS else 0
        for si in range(max(max_lunch, 1)):
            row = [""]
            if si == 0:
                row[0] = LUNCH_SERVER_TIME
            for day in DAYS:
                hibachi_lunch = self.schedule[day]["hibachi_lunch_servers"]
                servers = self.schedule[day]["lunch_servers"]
                letters = self.schedule[day].get("lunch_server_letters", {})
                b_set = set(self.schedule[day].get("b_side_lunch", []))
                hib_set = set(hibachi_lunch)
                combined = servers + hibachi_lunch
                combined.sort(key=lambda n: (
                    0 if n in b_set else 1,
                    letters.get(n, "Z"),
                    1 if n in hib_set else 0,
                ))
                mids = self.schedule[day].get("mid_servers", [])
                if si < len(combined):
                    name = combined[si]
                    letter = letters.get(name, chr(65 + si))
                    mid = "*" if name in mids else ""
                    if name in b_set:
                        row.append(f"<{name} {letter}>{mid}")
                    else:
                        row.append(f"{name} {letter}{mid}")
                else:
                    row.append("")
            y = draw_row(y, row, size=6)

        # Lunch manager
        row_mgr = ["Lunch Manager"]
        for day in DAYS:
            row_mgr.append(self.schedule[day]["lunch_manager"])
        y = draw_row(y, row_mgr, bold=True, size=6)
        # Lunch manager time
        row_mgr_t = [""] + [LUNCH_MANAGER_TIME] * 7
        y = draw_row(y, row_mgr_t, size=5)

        # Lunch hosts — name row then time row for each hostess slot
        max_lh = max(len(self.schedule[d]["lunch_hosts"]) for d in DAYS) if DAYS else 0
        for hi in range(max(max_lh, 1)):
            # Name row
            row = [f"Hostess {hi+1}" if hi < 3 else ""]
            for day in DAYS:
                hosts = self.schedule[day]["lunch_hosts"]
                row.append(hosts[hi] if hi < len(hosts) else "")
            y = draw_row(y, row, size=6)
            # Time row
            row_t = [""]
            for day in DAYS:
                hosts = self.schedule[day]["lunch_hosts"]
                if hi < len(hosts) and hi < len(LUNCH_HOST_TIMES):
                    row_t.append(LUNCH_HOST_TIMES[hi])
                else:
                    row_t.append("")
            y = draw_row(y, row_t, size=5)

        # Transition time row
        row_trans = [""]
        time_row_20 = {
            "MON": "(10:15-2:15)", "TUE": "(10:15-4:15)",
            "WED": "(4:00-8:00)", "THU": "(10:15-2:15)",
            "FRI": "(10:15-3:15)", "SAT": "(10:15-2:15)",
            "SUN": "(11:30-3:15)",
        }
        for day in DAYS:
            row_trans.append(time_row_20.get(day, ""))
        y = draw_row(y, row_trans, size=5)

        y += 1  # spacer

        # DINNER SECTION
        y = draw_row(y, ["DINNER SERVERS", "", "", "", "", "", "", ""], bold=True, size=7)

        max_dinner = max(len(self.schedule[d]["dinner_servers"]) + len(self.schedule[d]["hibachi_servers"]) for d in DAYS)
        for si in range(max(max_dinner, 1)):
            row = [""]
            if si == 0:
                row[0] = DINNER_SERVER_TIME
            for day in DAYS:
                hibachi = self.schedule[day]["hibachi_servers"]
                servers = self.schedule[day]["dinner_servers"]
                letters = self.schedule[day].get("dinner_server_letters", {})
                b_set = set(self.schedule[day].get("b_side_dinner", []))
                hib_set = set(hibachi)
                combined = servers + hibachi
                combined.sort(key=lambda n: (
                    0 if n in b_set else 1,
                    letters.get(n, "Z"),
                    1 if n in hib_set else 0,
                ))
                mids = self.schedule[day].get("mid_servers", [])
                if si < len(combined):
                    name = combined[si]
                    letter = letters.get(name, chr(65 + si))
                    mid = "*" if name in mids else ""
                    if name in b_set:
                        row.append(f"<{name} {letter}>{mid}")
                    else:
                        row.append(f"{name} {letter}{mid}")
                else:
                    row.append("")
            y = draw_row(y, row, size=6)

        # Dinner hosts — name row then time row for each hostess slot
        y += 1
        max_dh = max(len(self.schedule[d]["dinner_hosts"]) for d in DAYS) if DAYS else 0
        for hi in range(max(max_dh, 1)):
            # Name row
            row = [f"Hostess {hi+1}" if hi < 3 else ""]
            for day in DAYS:
                hosts = self.schedule[day]["dinner_hosts"]
                row.append(hosts[hi] if hi < len(hosts) else "")
            y = draw_row(y, row, size=6)
            # Time row
            row_t = [""]
            for di, day in enumerate(DAYS):
                hosts = self.schedule[day]["dinner_hosts"]
                if hi < len(hosts):
                    if di >= 5:  # SAT, SUN
                        times = DINNER_HOST_TIMES_WEEKEND if day != "SUN" else DINNER_HOST_TIMES_SUNDAY
                    elif di == 4:  # FRI
                        times = DINNER_HOST_TIMES_FRIDAY
                    else:
                        times = DINNER_HOST_TIMES_WEEKDAY
                    row_t.append(times[hi] if hi < len(times) else "")
                else:
                    row_t.append("")
            y = draw_row(y, row_t, size=5)

        # Dinner manager
        row_mgr = ["Dinner Manager"]
        for day in DAYS:
            row_mgr.append(self.schedule[day]["dinner_manager"])
        y = draw_row(y, row_mgr, bold=True, size=6)
        # Dinner manager time
        row_mgr_t = [""]
        for di, day in enumerate(DAYS):
            t = DINNER_MANAGER_TIME_WEEKEND if di >= 4 else DINNER_MANAGER_TIME_WEEKDAY
            row_mgr_t.append(t)
        y = draw_row(y, row_mgr_t, size=5)

        # Legend
        y += 3
        pdf.set_font("Helvetica", "I", 7)
        pdf.set_xy(x_start, y)
        pdf.cell(0, 4, "* = Midshift server(s)  |  Gold = Hibachi side  |  W/X/Y = B-side (sushi)  |  Letters = Sidework assignment")

        pdf.output(filepath)
        return filepath


# ---------------------------------------------------------------------------
# Main Application GUI
# ---------------------------------------------------------------------------
class _ExplorerGame:
    """2D Zelda-ish dungeon easter egg.

    Pitch-black room lit only by 7 torches. Enemies enter from the sides;
    space swings a sword that auto-aims at the nearest enemy with a small
    splash radius and a cooldown. Enemies take 3 hits, get knocked back, and
    have a 25% chance to drop a blue soul. Walk a soul onto a torch to
    extinguish it (you need its light less; the dungeon gets darker). Put
    out all 7 torches to win.
    """

    TILE = 32
    COLS = 30
    ROWS = 18
    HUD_HEIGHT = 38
    FRAME_MS = 90   # ~11 FPS

    SWORD_RANGE = 2.9     # mid-range reach
    SWORD_SPLASH = 1.4
    SWORD_COOLDOWN = 16   # was 7 — slower swing makes the game harder
    ENEMY_HP = 3
    ENEMY_SPEED = 0.085
    ENEMY_KNOCKBACK = 1.6
    ENEMY_KNOCKBACK_DECAY = 0.55
    ENEMY_DAMAGE_RANGE = 0.75  # how close before they hit you
    SPAWN_INTERVAL = 22  # frames between spawns
    SOUL_DROP_RATE = 0.25
    SOUL_DROP_RATE_LATE = 0.125   # halved chance for the final 2 souls
    SOUL_MAX_TOTAL = 7
    SOUL_LATE_THRESHOLD = 5       # after this many "in flight", switch to LATE rate
    SOUL_TTL_FRAMES = 111         # ~10 seconds at FRAME_MS=90
    NUM_TORCHES = 7
    POWERUP_DAMAGE_DROP_RATE = 0.125
    POWERUP_HEART_DROP_RATE = 0.125
    POWERUP_TTL_FRAMES = 167      # ~15 seconds before a dropped power-up vanishes
    DAMAGE_BOOST_FRAMES = 167     # ~15 seconds of 2x sword damage
    DAMAGE_BOOST_FRAMES_2X = 334  # ~30 seconds while two boosts are stacked
    DAMAGE_BOOST_FRAMES_BONUS = 56  # ~5 seconds added on overflow pickups
    DAMAGE_BOOST_MAX_STACKS = 2   # cap on simultaneous damage-boost stacks
    BOMB_RADIUS = 4.0             # tiles affected by the soul-bomb cheat
    BOMB_ANIM_FRAMES = 6
    PLAYER_MAX_HP = 3
    PLAYER_IFRAMES = 14
    VISIBILITY_THRESHOLD = 0.15  # enemies below this light intensity stay hidden

    # Hidden recruits unlocked by beating the game. Roles deliberately not
    # surfaced in the UI — the player picks blind.
    SHADOW_REALM_RECRUITS = [
        {"name": "DC",      "roles": ["server"],         "role_preference": "server"},
        {"name": "Thomas",  "roles": ["server"],         "role_preference": "server"},
        {"name": "Jessica", "roles": ["host"],           "role_preference": "host"},
        {"name": "Lynn",    "roles": ["server", "host"], "role_preference": None},
        {"name": "Johnny",  "roles": ["server", "host"], "role_preference": None},
        {"name": "Steven",  "roles": ["server", "host"], "role_preference": None},
        {"name": "Grace",   "roles": ["server", "host"], "role_preference": None},
    ]

    def __init__(self, parent, data=None, on_recruit=None):
        # Keep a reference to the launching window so the recruit popup can
        # be reparented to it after the game window is destroyed.
        self.parent = parent
        self.win = tk.Toplevel(parent)
        # DataManager — used by the win-screen recruit popup to add a new
        # staff member to the roster. Optional so the game can still run
        # standalone for testing.
        self.data = data
        # Callback fired after a recruit is appended + saved, so the App
        # can refresh its Staff treeview to show the new entry.
        self.on_recruit = on_recruit
        self.win.title("???")
        self.win.resizable(False, False)
        self.win.configure(bg="black")
        self.win.transient(parent)

        w = self.COLS * self.TILE
        h = self.ROWS * self.TILE + self.HUD_HEIGHT
        self.canvas = tk.Canvas(self.win, width=w, height=h,
                                bg="black", highlightthickness=0)
        self.canvas.pack()

        # No walls — open dungeon floor. Player and enemies are still kept in
        # bounds by the COLS/ROWS check in _move and _update_enemies.
        self.walls = set()

        # 7 torches scattered through the dungeon. Each is a dict so we can
        # toggle "lit" when extinguished by a soul.
        torch_positions = [
            (3, 2), (14, 2), (26, 2),
            (3, 15), (26, 15),
            (10, 9), (19, 9),
        ]
        self.torches = [{"x": x, "y": y, "lit": True} for (x, y) in torch_positions]

        # Player
        self.px, self.py = self.COLS // 2, self.ROWS - 3
        self.sword_cooldown = 0
        self.swing_anim = 0   # frames remaining of sword swing visual
        self.swing_target = None  # (cx, cy) of last swing
        # carried_souls is now a list of TTL counters (one per carried soul).
        # Souls expire from the inventory if not used in time.
        self.carried_souls = []
        self.hp = self.PLAYER_MAX_HP
        self.iframes = 0
        self.alive = True
        self.won = False
        self.game_over = False

        # Entities
        self.enemies = []   # {x, y, hp, kx, ky, hit_flash}
        self.souls = []     # {x, y, ttl}  — ttl in frames; despawn at 0
        self.power_ups = []  # {x, y, kind}  kind: "damage" | "heart"
        self.damage_boost = 0  # frames remaining of damage power-up
        self.damage_stacks = 0  # 0 = base, 1 = 2x, 2 = 3x (one-shot enemies)
        self.bomb_anim = 0  # frames remaining of soul-bomb explosion visual

        # Bindings — WASD only, plus space to attack
        for k, dx, dy in [
            ("w", 0, -1), ("a", -1, 0), ("s", 0, 1), ("d", 1, 0),
            ("W", 0, -1), ("A", -1, 0), ("S", 0, 1), ("D", 1, 0),
        ]:
            self.win.bind(k, lambda e, dx=dx, dy=dy: self._move(dx, dy))
        self.win.bind("<space>", lambda e: self._attack())
        # Hidden cheat: B consumes one carried soul as a mid-range bomb.
        self.win.bind("b", lambda e: self._bomb())
        self.win.bind("B", lambda e: self._bomb())
        self.win.bind("<Escape>", lambda e: self._close())
        self.win.protocol("WM_DELETE_WINDOW", self._close)
        self.win.focus_set()

        self._frame = 0
        self._spawn_counter = 0
        self._animate()

    # ---- input / actions ----

    def _close(self):
        self.alive = False
        try:
            self.win.destroy()
        except tk.TclError:
            pass

    def _move(self, dx, dy):
        if not self.alive or self.won:
            return
        nx, ny = self.px + dx, self.py + dy
        if (nx, ny) in self.walls:
            return
        if not (0 <= nx < self.COLS and 0 <= ny < self.ROWS):
            return
        self.px, self.py = nx, ny

        # Pick up any souls on this tile — each becomes a fresh inventory TTL
        keep = []
        for s in self.souls:
            if s["x"] == nx and s["y"] == ny:
                self.carried_souls.append(self.SOUL_TTL_FRAMES)
            else:
                keep.append(s)
        self.souls = keep

        # Pick up any power-ups on this tile. Hearts are only consumed if the
        # player actually has room to heal; damage boosts always apply.
        keep_pu = []
        for p in self.power_ups:
            if p["x"] == nx and p["y"] == ny:
                if p["kind"] == "damage":
                    # Stack up to DAMAGE_BOOST_MAX_STACKS. At 2 stacks the
                    # timer is reset AND extended by another 15s window
                    # (total 30s), rewarding the double pickup.
                    if self.damage_stacks < self.DAMAGE_BOOST_MAX_STACKS:
                        self.damage_stacks += 1
                        if self.damage_stacks >= 2:
                            self.damage_boost = self.DAMAGE_BOOST_FRAMES_2X
                        else:
                            self.damage_boost = self.DAMAGE_BOOST_FRAMES
                    else:
                        # Already at max stacks: refresh to 15s if the
                        # remaining timer is below that threshold,
                        # otherwise add a small +5s bonus on top, capped
                        # at the 30s 2x-stack ceiling.
                        if self.damage_boost < self.DAMAGE_BOOST_FRAMES:
                            self.damage_boost = self.DAMAGE_BOOST_FRAMES
                        else:
                            self.damage_boost = min(
                                self.DAMAGE_BOOST_FRAMES_2X,
                                self.damage_boost + self.DAMAGE_BOOST_FRAMES_BONUS,
                            )
                    continue
                elif p["kind"] == "heart":
                    if self.hp < self.PLAYER_MAX_HP:
                        self.hp += 1
                        continue
            keep_pu.append(p)
        self.power_ups = keep_pu

        # Feed a soul to a torch on this tile (if any, and we have one).
        # Use the oldest soul first (FIFO) so the timer pressure is fair.
        for t in self.torches:
            if t["x"] == nx and t["y"] == ny and t["lit"] and self.carried_souls:
                t["lit"] = False
                self.carried_souls.pop(0)
                if all(not tt["lit"] for tt in self.torches):
                    self._win()
                break

    def _attack(self):
        if not self.alive or self.won or self.game_over or self.sword_cooldown > 0:
            return
        # Cooldown applies on EVERY swing — even if it whiffs — so spamming
        # space when no enemy is around is genuinely punished.
        self.sword_cooldown = self.SWORD_COOLDOWN
        self.swing_anim = 2
        # Auto-aim: nearest enemy within range
        target = None
        best = self.SWORD_RANGE
        for e in self.enemies:
            d = math.hypot(e["x"] - self.px, e["y"] - self.py)
            if d <= best:
                best = d
                target = e
        if target is None:
            # Whiff — no damage, but the swing still played and cooldown is set.
            # Mark target as None so the draw code shows a miss arc instead.
            self.swing_target = None
            return
        self.swing_target = (target["x"], target["y"])
        # Apply damage + knockback to target and anyone in splash radius.
        # Damage scales with damage-boost stacks: 0 → 1, 1 → 2 (2x),
        # 2 → 3 (3x, one-shots ENEMY_HP=3 enemies).
        sword_damage = 1 + self.damage_stacks if self.damage_boost > 0 else 1
        for e in self.enemies:
            d = math.hypot(e["x"] - target["x"], e["y"] - target["y"])
            if d <= self.SWORD_SPLASH:
                e["hp"] -= sword_damage
                kdx = e["x"] - self.px
                kdy = e["y"] - self.py
                kd = math.hypot(kdx, kdy)
                if kd > 0.01:
                    kdx /= kd
                    kdy /= kd
                else:
                    kdx, kdy = 0, -1
                e["kx"] = kdx * self.ENEMY_KNOCKBACK
                e["ky"] = kdy * self.ENEMY_KNOCKBACK
                e["hit_flash"] = 3
        self._process_enemy_deaths()

    def _clamp_drop_to_corner(self, x, y):
        """Return an in-bounds tile for a drop. If (x, y) is already on the
        playfield it's returned unchanged; otherwise the drop is redirected
        to the corner of the playfield nearest to (x, y) so the player can
        still walk over and collect it."""
        if 0 <= x < self.COLS and 0 <= y < self.ROWS:
            return x, y
        cx = 0 if x < self.COLS / 2 else self.COLS - 1
        cy = 0 if y < self.ROWS / 2 else self.ROWS - 1
        return cx, cy

    def _process_enemy_deaths(self):
        """Remove dead enemies, roll soul/power-up drops on each kill.

        Soul drop cap: extinguished torches + on-ground + carried < 7.
        Last 2 souls drop at half rate. Power-ups roll independently."""
        survivors = []
        for e in self.enemies:
            if e["hp"] <= 0:
                ex_int = int(round(e["x"]))
                ey_int = int(round(e["y"]))
                # If knockback launched the enemy off the playfield, the
                # rounded drop tile may end up outside the grid (or in the
                # HUD strip), making the drop unreachable. Redirect those
                # drops to the in-bounds corner nearest where the enemy
                # ended up so the player can still pick them up.
                ex_int, ey_int = self._clamp_drop_to_corner(ex_int, ey_int)
                used = sum(1 for t in self.torches if not t["lit"])
                in_flight = len(self.souls) + len(self.carried_souls)
                accounted = used + in_flight
                if accounted < self.SOUL_MAX_TOTAL:
                    rate = (self.SOUL_DROP_RATE_LATE
                            if accounted >= self.SOUL_LATE_THRESHOLD
                            else self.SOUL_DROP_RATE)
                    if random.random() < rate:
                        self.souls.append({
                            "x": ex_int, "y": ey_int,
                            "ttl": self.SOUL_TTL_FRAMES,
                        })
                if random.random() < self.POWERUP_DAMAGE_DROP_RATE:
                    self.power_ups.append({
                        "x": ex_int, "y": ey_int, "kind": "damage",
                        "ttl": self.POWERUP_TTL_FRAMES,
                    })
                if random.random() < self.POWERUP_HEART_DROP_RATE:
                    self.power_ups.append({
                        "x": ex_int, "y": ey_int, "kind": "heart",
                        "ttl": self.POWERUP_TTL_FRAMES,
                    })
            else:
                survivors.append(e)
        self.enemies = survivors

    def _bomb(self):
        """Hidden cheat: consume the oldest carried soul as a bomb. Mid-range
        blue splash damage equal to two sword swings. The soul does NOT count
        toward the 7 needed — popping it from the inventory frees the cap so
        an enemy can drop a replacement."""
        if not self.alive or self.won or self.game_over:
            return
        if not self.carried_souls:
            return
        self.carried_souls.pop(0)
        self.bomb_anim = self.BOMB_ANIM_FRAMES
        sword_damage = 1 + self.damage_stacks if self.damage_boost > 0 else 1
        bomb_damage = sword_damage * 2  # "2 sword swings worth"
        for e in self.enemies:
            d = math.hypot(e["x"] - self.px, e["y"] - self.py)
            if d <= self.BOMB_RADIUS:
                e["hp"] -= bomb_damage
                kdx = e["x"] - self.px
                kdy = e["y"] - self.py
                kd = math.hypot(kdx, kdy)
                if kd > 0.01:
                    kdx /= kd
                    kdy /= kd
                else:
                    kdx, kdy = 0, -1
                e["kx"] = kdx * self.ENEMY_KNOCKBACK * 1.6
                e["ky"] = kdy * self.ENEMY_KNOCKBACK * 1.6
                e["hit_flash"] = 4
        self._process_enemy_deaths()

    # ---- entity update ----

    def _spawn_enemy(self):
        # Enter from the left or right edge
        side = random.choice(("left", "right"))
        x = 1 if side == "left" else self.COLS - 2
        y = random.randint(2, self.ROWS - 3)
        if (x, y) in self.walls:
            return
        self.enemies.append({
            "x": float(x), "y": float(y),
            "hp": self.ENEMY_HP,
            "kx": 0.0, "ky": 0.0,
            "hit_flash": 0,
        })

    def _update_enemies(self):
        for e in self.enemies:
            # Knockback first (overrides AI movement while it lasts)
            if abs(e["kx"]) > 0.05 or abs(e["ky"]) > 0.05:
                nx = e["x"] + e["kx"]
                ny = e["y"] + e["ky"]
                if (0 <= nx < self.COLS and 0 <= ny < self.ROWS
                        and (int(round(nx)), int(round(ny))) not in self.walls):
                    e["x"] = nx
                    e["y"] = ny
                e["kx"] *= self.ENEMY_KNOCKBACK_DECAY
                e["ky"] *= self.ENEMY_KNOCKBACK_DECAY
            else:
                # Walk toward player
                dx = self.px - e["x"]
                dy = self.py - e["y"]
                d = math.hypot(dx, dy)
                if d > 0.4:
                    nx = e["x"] + dx / d * self.ENEMY_SPEED
                    ny = e["y"] + dy / d * self.ENEMY_SPEED
                    if (int(round(nx)), int(round(ny))) not in self.walls:
                        e["x"] = nx
                        e["y"] = ny
            if e["hit_flash"] > 0:
                e["hit_flash"] -= 1

        # Damage check: any enemy within ENEMY_DAMAGE_RANGE of player
        # deals 1 damage if iframes are 0.
        if self.iframes == 0 and self.hp > 0 and not self.game_over:
            for e in self.enemies:
                d = math.hypot(e["x"] - self.px, e["y"] - self.py)
                if d < self.ENEMY_DAMAGE_RANGE:
                    self.hp -= 1
                    self.iframes = self.PLAYER_IFRAMES
                    if self.hp <= 0:
                        self._trigger_game_over()
                    break

    # ---- frame loop ----

    def _animate(self):
        if not self.alive:
            return
        if self.won or self.game_over:
            return
        self._frame += 1
        if self.sword_cooldown > 0:
            self.sword_cooldown -= 1
        if self.swing_anim > 0:
            self.swing_anim -= 1
        if self.iframes > 0:
            self.iframes -= 1
        if self.damage_boost > 0:
            self.damage_boost -= 1
            if self.damage_boost == 0:
                self.damage_stacks = 0
        if self.bomb_anim > 0:
            self.bomb_anim -= 1
        # Tick soul TTLs (ground + carried) and prune expired
        for s in self.souls:
            s["ttl"] -= 1
        self.souls = [s for s in self.souls if s["ttl"] > 0]
        self.carried_souls = [t - 1 for t in self.carried_souls if t - 1 > 0]
        # Tick power-up TTLs and prune expired
        for p in self.power_ups:
            p["ttl"] = p.get("ttl", self.POWERUP_TTL_FRAMES) - 1
        self.power_ups = [p for p in self.power_ups if p["ttl"] > 0]
        self._spawn_counter += 1
        if self._spawn_counter >= self.SPAWN_INTERVAL:
            self._spawn_counter = 0
            self._spawn_enemy()
        self._update_enemies()
        # _update_enemies may have triggered game over — draw the GO screen
        # AFTER the regular game frame would overwrite it. (Previously the
        # immediate draw inside _trigger_game_over got clobbered by self._draw.)
        if self.game_over:
            self._draw_game_over_screen()
            return
        self._draw()
        try:
            self.win.after(self.FRAME_MS, self._animate)
        except tk.TclError:
            pass

    def _win(self):
        self.won = True
        self._draw_win_screen()
        # Let the win screen sit for a moment, then close the game window
        # entirely. The recruit popup appears AFTER the game closes, so it
        # feels like a reward delivered back in the main app.
        self.win.after(5000, self._close_then_show_recruit)

    def _close_then_show_recruit(self):
        try:
            self.win.destroy()
        except tk.TclError:
            pass
        self.alive = False
        # Schedule the recruit popup on the original parent so it survives
        # the game window being destroyed.
        try:
            self.parent.after(250, self._show_recruit_popup)
        except tk.TclError:
            pass

    def _show_recruit_popup(self):
        """Win-screen recruit picker. Lets the player add ONE Shadow Realm
        staff member to the roster. Roles are intentionally hidden — the
        dropdown shows names only. Bound to BackSpace as a debug shortcut."""
        if self.data is None:
            self._close()
            return
        # The popup is parented to the original launching window because by
        # the time we get here the game window has already been destroyed.
        popup_parent = self.parent if self.parent is not None else self.win
        # Filter out anyone already on the roster so the player can't add a
        # duplicate name (would error out the staff editor).
        existing = {s["name"] for s in self.data.staff}
        available = [r for r in self.SHADOW_REALM_RECRUITS if r["name"] not in existing]
        if not available:
            messagebox.showinfo(
                "The Shadow Realm",
                "All Shadow Realm souls have already joined Toyo.",
                parent=popup_parent)
            self._close()
            return

        popup = tk.Toplevel(popup_parent)
        popup.title("The Shadow Realm")
        popup.geometry("440x220")
        popup.configure(bg="black")
        popup.resizable(False, False)
        popup.transient(popup_parent)
        popup.after(50, lambda: popup.grab_set() if popup.winfo_exists() else None)

        tk.Label(popup,
                 text="Who would you like to recruit",
                 font=("Helvetica", 13, "bold"),
                 bg="black", fg="white").pack(pady=(22, 0))
        tk.Label(popup,
                 text="from The Shadow Realm?",
                 font=("Helvetica", 13, "bold"),
                 bg="black", fg="#FFD54F").pack(pady=(0, 14))

        # "None" lets the player decline a recruit if they'd rather not
        # add anyone from the Shadow Realm to the roster.
        NONE_LABEL = "None — don't recruit anyone"
        name_var = tk.StringVar(value=available[0]["name"])
        combo = ttk.Combobox(popup, textvariable=name_var,
                             values=[r["name"] for r in available] + [NONE_LABEL],
                             state="readonly", width=26,
                             font=("Helvetica", 12))
        combo.pack(pady=4)

        def confirm():
            chosen_name = name_var.get()
            if chosen_name == NONE_LABEL:
                try:
                    popup.destroy()
                except tk.TclError:
                    pass
                self._close()
                return
            chosen = next((r for r in available if r["name"] == chosen_name), None)
            if chosen:
                new_staff = {
                    "name": chosen["name"],
                    "roles": list(chosen["roles"]),
                    "seniority": 5,
                    "fixed_schedule": False,
                    "active": True,
                    # Exclusive flag awarded by clearing the easter egg.
                    # Visual-only — no scheduling effect — and the staff
                    # dialog cannot set or unset it.
                    "flags": ["ABSOLUTE EMERGENCY"],
                    "role_preference": chosen["role_preference"],
                    "default_off": [],
                }
                self.data.staff.append(new_staff)
                try:
                    self.data.save_staff()
                except Exception:
                    pass
                # Notify the App so the Staff treeview picks up the new entry
                if self.on_recruit is not None:
                    try:
                        self.on_recruit(new_staff)
                    except Exception:
                        pass
            try:
                popup.destroy()
            except tk.TclError:
                pass
            self._close()

        tk.Button(popup, text="Recruit",
                  font=("Helvetica", 11, "bold"),
                  bg="#4CAF50", fg="white",
                  activebackground="#388E3C", activeforeground="white",
                  relief="raised", bd=2, cursor="hand2",
                  command=confirm).pack(pady=18, ipadx=14, ipady=4)

        popup.protocol("WM_DELETE_WINDOW",
                       lambda: (popup.destroy(), self._close()))

    def _trigger_game_over(self):
        self.game_over = True
        # The actual draw happens at the end of the current _animate tick,
        # so the regular _draw() call doesn't overwrite the GO screen.
        self.win.after(5000, self._close)

    def _draw_game_over_screen(self):
        c = self.canvas
        c.delete("all")
        w = self.COLS * self.TILE
        h = self.ROWS * self.TILE + self.HUD_HEIGHT
        c.create_rectangle(0, 0, w, h, fill="black", outline="")
        c.create_text(w / 2, h / 2 - 26,
                      text="GAME OVER",
                      fill="#FF4444", font=("Courier", 32, "bold"))
        c.create_text(w / 2, h / 2 + 24,
                      text="The Toyo crew is gone...",
                      fill="white", font=("Courier", 14, "bold"))

    # 8-bit pixel art of the Philosopher's Stone — a red diamond gem with
    # white inner highlights and a darker outer rim. '.' = transparent.
    PHILOSOPHER_STONE_ART = [
        ".....R.....",
        "....RRR....",
        "...RRrRR...",
        "..RRrwrRR..",
        ".RRrwwwrRR.",
        "RRrwwwwwrRR",
        ".RRrwwwrRR.",
        "..RRrwrRR..",
        "...RRrRR...",
        "....RRR....",
        ".....R.....",
    ]
    PHILOSOPHER_STONE_PALETTE = {
        "R": "#5A0000",
        "r": "#D32F2F",
        "w": "#FFFFFF",
    }

    def _draw_philosopher_stone(self, cx, cy, pixel=10):
        """Render the 8-bit philosopher stone centered on (cx, cy)."""
        c = self.canvas
        rows = self.PHILOSOPHER_STONE_ART
        h = len(rows)
        w = len(rows[0])
        x0 = cx - (w * pixel) // 2
        y0 = cy - (h * pixel) // 2
        for ry, row in enumerate(rows):
            for rx, ch in enumerate(row):
                color = self.PHILOSOPHER_STONE_PALETTE.get(ch)
                if color is None:
                    continue
                px0 = x0 + rx * pixel
                py0 = y0 + ry * pixel
                c.create_rectangle(px0, py0, px0 + pixel, py0 + pixel,
                                   fill=color, outline="")

    def _draw_win_screen(self):
        c = self.canvas
        c.delete("all")
        w = self.COLS * self.TILE
        h = self.ROWS * self.TILE + self.HUD_HEIGHT
        c.create_rectangle(0, 0, w, h, fill="black", outline="")
        # Pixel art philosopher's stone above the text
        self._draw_philosopher_stone(w / 2, h / 2 - 80, pixel=12)
        c.create_text(w / 2, h / 2 + 60,
                      text="You have collected all 7 souls.",
                      fill="white", font=("Courier", 18, "bold"))
        c.create_text(w / 2, h / 2 + 100,
                      text="You have obtained The Philosopher Stone.",
                      fill="white", font=("Courier", 18, "bold"))

    # ---- rendering ----

    def _draw(self):
        c = self.canvas
        c.delete("all")
        T = self.TILE

        # Build per-frame light source list. Each entry: (lx, ly, lr, color)
        # color = "warm" for player + torches, "blue" for souls.
        sources = []
        sources.append((self.px + 0.5, self.py + 0.5,
                        2.9 + random.uniform(-0.12, 0.12), "warm"))
        for ti, t in enumerate(self.torches):
            if not t["lit"]:
                continue
            base = 4.4
            jitter = (math.sin(self._frame * 0.5 + ti * 1.3) * 0.22
                      + random.uniform(-0.3, 0.3))
            sources.append((t["x"] + 0.5, t["y"] + 0.5, base + jitter, "warm"))
        for s in self.souls:
            # Fresh souls cast a much brighter light; the radius shrinks as
            # the TTL ticks down so dying souls only barely glow.
            ttl_ratio = max(0.0, s["ttl"] / self.SOUL_TTL_FRAMES)
            sr = 1.0 + 2.6 * ttl_ratio + random.uniform(-0.15, 0.15)
            sources.append((s["x"] + 0.5, s["y"] + 0.5, sr, "blue"))

        # Bomb flash — while the bomb anim is playing, drop a huge blue
        # light source on the player so the surrounding tiles brighten in
        # cool blue. This is the "illuminates the darkness" effect.
        if self.bomb_anim > 0:
            flash_progress = self.bomb_anim / self.BOMB_ANIM_FRAMES
            flash_radius = self.BOMB_RADIUS * 1.5 * (0.5 + 0.5 * flash_progress)
            sources.append((self.px + 0.5, self.py + 0.5, flash_radius, "blue"))

        # Per-tile lighting: pick the strongest contribution and use its color.
        for y in range(self.ROWS):
            for x in range(self.COLS):
                cx, cy = x + 0.5, y + 0.5
                best_i = 0.0
                best_color = "warm"
                for lx, ly, lr, ct in sources:
                    d = math.hypot(cx - lx, cy - ly)
                    if d < lr:
                        i = (1.0 - d / lr) ** 1.4
                        if i > best_i:
                            best_i = i
                            best_color = ct
                if best_i < 0.04:
                    continue
                is_wall = (x, y) in self.walls
                fill = self._tile_color(best_i, is_wall, best_color)
                c.create_rectangle(x * T, y * T,
                                   (x + 1) * T, (y + 1) * T,
                                   fill=fill, outline="")

        # Lit torches
        for t in self.torches:
            if t["lit"]:
                self._draw_flame(t["x"] * T + T // 2, t["y"] * T + T // 2 + 2,
                                 hue="warm")

        # Souls on the ground (blue dynamic fires) — flame size also scales
        # with remaining TTL so a dying soul shrinks visibly.
        for s in self.souls:
            ttl_ratio = max(0.05, s["ttl"] / self.SOUL_TTL_FRAMES)
            scale = 0.4 + 0.7 * ttl_ratio
            self._draw_flame(s["x"] * T + T // 2, s["y"] * T + T // 2 + 2,
                             hue="blue", scale=scale)

        # Power-ups on the ground — only visible if their tile is lit above
        # the visibility threshold (same rule as enemies). Walking onto an
        # unseen power-up still picks it up; you just can't see it.
        pulse = 1.0 + math.sin(self._frame * 0.4) * 0.15
        for p in self.power_ups:
            pcx = p["x"] + 0.5
            pcy = p["y"] + 0.5
            best = 0.0
            for lx, ly, lr, _ in sources:
                d = math.hypot(pcx - lx, pcy - ly)
                if d < lr:
                    i = (1.0 - d / lr) ** 1.4
                    if i > best:
                        best = i
            if best < self.VISIBILITY_THRESHOLD:
                continue
            cx = p["x"] * T + T // 2
            cy = p["y"] * T + T // 2
            if p["kind"] == "damage":
                size = int(9 * pulse)
                pts = [cx, cy - size, cx + size, cy,
                       cx, cy + size, cx - size, cy]
                c.create_polygon(pts, fill="#FFD700", outline="#FFF59D", width=2)
                c.create_text(cx, cy, text="2x",
                              fill="#5A4400", font=("Helvetica", 7, "bold"))
            elif p["kind"] == "heart":
                size = int(9 * pulse)
                pts = [
                    cx, cy + size,
                    cx + size, cy,
                    cx + int(size * 0.55), cy - int(size * 0.7),
                    cx, cy - int(size * 0.15),
                    cx - int(size * 0.55), cy - int(size * 0.7),
                    cx - size, cy,
                ]
                c.create_polygon(pts, fill="#FF3B3B",
                                 outline="#FFCDD2", width=2)

        # Enemies — only visible if their tile is lit above the threshold.
        # Hit-flash enemies are always shown briefly so feedback is readable.
        for e in self.enemies:
            cx, cy = e["x"] + 0.5, e["y"] + 0.5
            best = 0.0
            for lx, ly, lr, _ in sources:
                d = math.hypot(cx - lx, cy - ly)
                if d < lr:
                    i = (1.0 - d / lr) ** 1.4
                    if i > best:
                        best = i
            if best < self.VISIBILITY_THRESHOLD and e["hit_flash"] == 0:
                continue
            ex = e["x"] * T + T // 2
            ey = e["y"] * T + T // 2
            half = T // 2 - 7
            if e["hit_flash"] > 0:
                fill, outline = "#ffffff", "#ff8a80"
            else:
                fill, outline = "#9c1a1a", "#ff5252"
            c.create_rectangle(ex - half, ey - half, ex + half, ey + half,
                               fill=fill, outline=outline, width=2)

        # Sword swing visual
        if self.swing_anim > 0:
            x1 = self.px * T + T // 2
            y1 = self.py * T + T // 2
            if self.swing_target is not None:
                # Hit: line to target + yellow splash circle
                tx, ty = self.swing_target
                x2 = tx * T + T // 2
                y2 = ty * T + T // 2
                c.create_line(x1, y1, x2, y2,
                              fill="#FFFFFF", width=4)
                r = self.SWORD_SPLASH * T
                c.create_oval(x2 - r, y2 - r, x2 + r, y2 + r,
                              outline="#FFE082", width=2)
            else:
                # Miss: dim grey arc around the player showing the range
                r = self.SWORD_RANGE * T
                c.create_oval(x1 - r, y1 - r, x1 + r, y1 + r,
                              outline="#9e9e9e", width=2, dash=(3, 3))

        # Bomb splash: layered concentric blue rings that fade as the anim
        # ends. The huge blue light source above already brightens the area;
        # this is the visible "splash" overlay.
        if self.bomb_anim > 0:
            cx = self.px * T + T // 2
            cy = self.py * T + T // 2
            fade = self.bomb_anim / self.BOMB_ANIM_FRAMES  # 1.0 → 0.0
            base_r = self.BOMB_RADIUS * T
            for mult, color, width in [
                (1.00, "#0d47a1", 2),
                (0.78, "#1976d2", 3),
                (0.58, "#2196f3", 3),
                (0.40, "#64b5f6", 4),
                (0.24, "#bbdefb", 4),
            ]:
                r = base_r * mult * (0.6 + 0.4 * fade)
                c.create_oval(cx - r, cy - r, cx + r, cy + r,
                              outline=color, width=width)

        # Player (orange square) — flickers during invincibility frames
        if self.iframes == 0 or (self.iframes // 2) % 2 == 0:
            px = self.px * T + 5
            py = self.py * T + 5
            c.create_rectangle(px, py, px + T - 10, py + T - 10,
                               fill="#FF8C00", outline="#FFB347", width=2)

        # HUD bar — minimal: hearts and any carried soul dots.
        hud_y = self.ROWS * T
        c.create_rectangle(0, hud_y, self.COLS * T, hud_y + self.HUD_HEIGHT,
                           fill="#0a0a14", outline="#1f1f38")
        hearts = "♥" * self.hp + "♡" * (self.PLAYER_MAX_HP - self.hp)
        c.create_text(
            14, hud_y + self.HUD_HEIGHT // 2,
            text=f"  {hearts}",
            fill="#FF6B6B", font=("Helvetica", 16, "bold"), anchor="w")

        # Carried soul dots — one fading blue circle per carried soul, drawn
        # right of the hearts. Brightness scales with remaining TTL so the
        # player can see how close each soul is to vanishing.
        cy_dot = hud_y + self.HUD_HEIGHT // 2
        x_dot = 100
        for ttl in self.carried_souls:
            ratio = max(0.0, min(1.0, ttl / self.SOUL_TTL_FRAMES))
            r = int(20 + 80 * ratio)
            g = int(60 + 120 * ratio)
            b = int(120 + 135 * ratio)
            color = f"#{r:02x}{g:02x}{b:02x}"
            c.create_oval(x_dot - 7, cy_dot - 7, x_dot + 7, cy_dot + 7,
                          fill=color, outline="")
            x_dot += 20

        # Damage-boost indicator: yellow diamond labelled with the active
        # multiplier. 1 stack = 2x, 2 stacks = 3x (one-shots enemies).
        if self.damage_boost > 0:
            x_dot += 6
            size = 10
            mult_label = f"{1 + self.damage_stacks}x"
            # Brighter outline at 2 stacks so the boosted state is obvious.
            outline = "#FFFFFF" if self.damage_stacks >= 2 else "#FFF59D"
            pts = [x_dot, cy_dot - size, x_dot + size, cy_dot,
                   x_dot, cy_dot + size, x_dot - size, cy_dot]
            c.create_polygon(pts, fill="#FFD700",
                             outline=outline, width=2)
            c.create_text(x_dot, cy_dot, text=mult_label,
                          fill="#5A4400", font=("Helvetica", 8, "bold"))

    @staticmethod
    def _tile_color(intensity, is_wall, color="warm"):
        """Tint a tile based on light intensity (0..1) and source color."""
        intensity = max(0.0, min(1.0, intensity))
        if color == "blue":
            if is_wall:
                r = int(15 + intensity * 50)
                g = int(25 + intensity * 75)
                b = int(50 + intensity * 150)
            else:
                r = int(5 + intensity * 30)
                g = int(15 + intensity * 60)
                b = int(30 + intensity * 130)
        else:  # warm
            if is_wall:
                r = int(40 + intensity * 130)
                g = int(20 + intensity * 65)
                b = int(15 + intensity * 35)
            else:
                r = int(15 + intensity * 110)
                g = int(8 + intensity * 55)
                b = int(5 + intensity * 25)
        return f"#{r:02x}{g:02x}{b:02x}"

    def _draw_flame(self, cx, cy, hue="warm", scale=1.0):
        """Procedural flickering flame. Smaller layered ovals than before."""
        c = self.canvas
        if hue == "warm":
            layers = [
                (8, "#3a0a00"),
                (6, "#7a1500"),
                (4, "#cc3300"),
                (3, "#ff6600"),
                (2, "#ffa726"),
                (1, "#fff176"),
            ]
        else:  # blue souls
            layers = [
                (8, "#001a3a"),
                (6, "#00337a"),
                (4, "#0066cc"),
                (3, "#0099ff"),
                (2, "#33ccff"),
                (1, "#ccf2ff"),
            ]
        for radius, color in layers:
            jx = random.randint(-2, 2)
            jy = random.randint(-2, 1)
            rj = random.uniform(-1.0, 1.0)
            r = max(1, (radius + rj) * scale)
            c.create_oval(cx - r + jx, cy - r + jy,
                          cx + r + jx, cy + r + jy,
                          fill=color, outline="")


class ToyoSchedulerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Toyo Schedule Maker (Light)")
        self.root.geometry("1200x750")
        self.root.minsize(1000, 600)

        self.data = DataManager()
        self.generator = ScheduleGenerator(self.data)
        self.current_schedule = None
        self.current_warnings = []

        # Tracks the single currently-open +Add popup on the Availability tab
        # so we can enforce one-at-a-time and toggle on repeat clicks.
        self._active_add_popup = None
        self._active_add_btn = None

        style = ttk.Style()
        style.configure("TNotebook.Tab", padding=[14, 8], font=("Helvetica", 11))
        style.configure("Treeview", rowheight=32, font=("Helvetica", 12))
        style.configure("Treeview.Heading", font=("Helvetica", 13, "bold"))
        # Reusable larger button style for the main tab toolbars
        style.configure("Big.TButton", font=("Helvetica", 12))

        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        # Close any open +Add popup when the user switches tabs
        self.notebook.bind("<<NotebookTabChanged>>", self._close_active_add_popup)

        self._build_staff_tab()
        self._build_config_tab()
        self._build_availability_tab()
        self._build_schedule_tab()
        self._build_export_tab()

    def _close_active_add_popup(self, event=None):
        """Destroy the currently-open +Add popup, if any. Called on tab switch."""
        if self._active_add_popup is not None:
            try:
                self._active_add_popup.destroy()
            except tk.TclError:
                pass
            self._active_add_popup = None
            self._active_add_btn = None

    # ---- TAB 1: STAFF MANAGEMENT ----
    def _build_staff_tab(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="  Staff  ")

        # Toolbar
        toolbar = ttk.Frame(frame)
        toolbar.pack(fill=tk.X, padx=8, pady=8)
        ttk.Button(toolbar, text="Add Staff", command=self._add_staff,
                   style="Big.TButton").pack(side=tk.LEFT, padx=3, ipadx=6, ipady=3)
        ttk.Button(toolbar, text="Edit Selected", command=self._edit_staff,
                   style="Big.TButton").pack(side=tk.LEFT, padx=3, ipadx=6, ipady=3)
        ttk.Button(toolbar, text="Remove Selected", command=self._remove_staff,
                   style="Big.TButton").pack(side=tk.LEFT, padx=3, ipadx=6, ipady=3)
        ttk.Button(toolbar, text="Reset to Defaults", command=self._reset_staff,
                   style="Big.TButton").pack(side=tk.RIGHT, padx=3, ipadx=6, ipady=3)
        tk.Button(toolbar, text="❓  Staff Help",
                  font=("Helvetica", 11, "bold"),
                  bg="#FFD54F", fg="black", activebackground="#FFC107",
                  relief="raised", bd=2, cursor="hand2",
                  command=lambda: self._show_flags_help(self.root)).pack(
            side=tk.RIGHT, padx=8, ipadx=6, ipady=3)

        # Treeview
        cols = ("Name", "Roles", "Seniority", "Preference", "Flags", "Active")
        self.staff_tree = ttk.Treeview(frame, columns=cols, show="headings", height=18)
        col_widths = {
            "Name": 160, "Roles": 130, "Seniority": 110,
            "Preference": 130, "Flags": 240, "Active": 90,
        }
        for col in cols:
            self.staff_tree.heading(col, text=col)
            self.staff_tree.column(col, width=col_widths[col], anchor="center")

        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.staff_tree.yview)
        self.staff_tree.configure(yscrollcommand=scrollbar.set)
        self.staff_tree.pack(fill=tk.BOTH, expand=True, padx=5, side=tk.LEFT)
        scrollbar.pack(fill=tk.Y, side=tk.LEFT)

        self.staff_tree.bind("<Double-1>", lambda e: self._edit_staff())
        self._refresh_staff_tree()

    def _refresh_staff_tree(self):
        for item in self.staff_tree.get_children():
            self.staff_tree.delete(item)
        for s in self.data.staff:
            roles = ", ".join(s["roles"])
            flags = ", ".join(s.get("flags", []))
            pref = s.get("role_preference", "") or ""
            active = "Yes" if s["active"] else "No"
            self.staff_tree.insert("", "end", values=(
                s["name"], roles, s["seniority"], pref, flags, active))

    def _add_staff(self):
        self._staff_dialog(None)

    def _edit_staff(self):
        sel = self.staff_tree.selection()
        if not sel:
            messagebox.showinfo("Select", "Select a staff member to edit.")
            return
        values = self.staff_tree.item(sel[0])["values"]
        name = values[0]
        staff = self.data.get_staff_by_name(name)
        if staff:
            self._staff_dialog(staff)

    def _remove_staff(self):
        sel = self.staff_tree.selection()
        if not sel:
            return
        values = self.staff_tree.item(sel[0])["values"]
        name = values[0]
        if messagebox.askyesno("Remove", f"Remove {name} from roster?"):
            self.data.staff = [s for s in self.data.staff if s["name"] != name]
            self.data.save_staff()
            self._refresh_staff_tree()

    def _reset_staff(self):
        if messagebox.askyesno("Reset", "Reset staff to factory defaults?"):
            self.data.staff = copy.deepcopy(DEFAULT_STAFF)
            self.data.save_staff()
            self._refresh_staff_tree()

    def _show_flags_help(self, parent):
        """Popup explaining what each staff flag and the Day Offs section do."""
        flags_doc = [
            ("fixed_schedule",
             "Staff member has a preset weekly schedule that overrides normal "
             "availability. Used for managers (Aaron, Chan) and fixed-shift "
             "servers like Dian and Leony."),
            ("always_hibachi",
             "This server is always assigned to hibachi shifts when working. "
             "Used for Will."),
            ("no_closing",
             "Staff member cannot work the closing host shift. Used for "
             "people who must leave early (e.g. Maria)."),
            ("fill_in",
             "Not used in the primary scheduling pass, but the generator will "
             "pull them in as a backup if there aren't enough regular servers "
             "to cover a shift."),
            ("emergency_only",
             "Never auto-scheduled, even as a backup. Only assignable "
             "manually through the schedule edit dropdown. Used for staff "
             "like Winnie, Ross, and Shayne."),
            ("seniority_priority",
             "Staff member gets priority shift assignments based on their "
             "seniority score. Used for Olivia (host)."),
        ]
        sections_doc = [
            ("Day Offs",
             "Recurring days a staff member doesn't work. Currently this "
             "field is only enforced for fulltimers (staff with the "
             "fixed_schedule flag) — Aaron, Chan, Dian, Leony, and Olivia. "
             "The scheduler uses it as the off-day fallback for fulltimers "
             "when no explicit availability is set, and the 'Set Default' "
             "button on the Availability tab restores fulltimers to their "
             "off-days. For non-fulltimer staff, the field is saved but "
             "ignored — their availability comes entirely from the "
             "Availability tab."),
        ]

        win = tk.Toplevel(parent)
        win.title("Staff Help")
        win.geometry("560x560")
        win.transient(parent)
        win.grab_set()

        ttk.Label(win, text="Staff Help",
                  font=("Helvetica", 14, "bold")).pack(pady=(12, 4))

        # Scrollable body so future entries don't overflow
        outer = tk.Frame(win, bd=1, relief="sunken")
        outer.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)
        canvas = tk.Canvas(outer, bg="white", highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        body = tk.Frame(canvas, bg="white")
        body.bind("<Configure>",
                  lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        tk.Label(body, text="Staff Flags",
                 font=("Helvetica", 12, "bold"),
                 bg="white", fg="#333333", anchor="w").pack(
            anchor="w", padx=10, pady=(8, 2))
        for name, desc in flags_doc:
            row = tk.Frame(body, bg="white")
            row.pack(fill=tk.X, padx=10, pady=6, anchor="w")
            tk.Label(row, text=name, font=("Helvetica", 11, "bold"),
                     fg="#0066CC", bg="white", anchor="w").pack(anchor="w")
            tk.Label(row, text=desc, font=("Helvetica", 10),
                     bg="white", anchor="w", justify="left",
                     wraplength=500).pack(anchor="w", padx=(12, 0))

        # Other dialog sections (e.g. Day Offs)
        tk.Label(body, text="Other Sections",
                 font=("Helvetica", 12, "bold"),
                 bg="white", fg="#333333", anchor="w").pack(
            anchor="w", padx=10, pady=(14, 2))
        for name, desc in sections_doc:
            row = tk.Frame(body, bg="white")
            row.pack(fill=tk.X, padx=10, pady=6, anchor="w")
            tk.Label(row, text=name, font=("Helvetica", 11, "bold"),
                     fg="#0066CC", bg="white", anchor="w").pack(anchor="w")
            tk.Label(row, text=desc, font=("Helvetica", 10),
                     bg="white", anchor="w", justify="left",
                     wraplength=500).pack(anchor="w", padx=(12, 0))

        ttk.Button(win, text="Close", command=win.destroy).pack(pady=10, ipadx=10)

    def _absolute_emergency(self, parent):
        """Easter egg flow: confirmation → 'Very well then...' → loading
        screen → mini-game. Triggered by the ABSOLUTE EMERGENCY button in
        the staff dialog."""
        if not messagebox.askyesno(
                "ABSOLUTE EMERGENCY",
                "Are you sure?\n\n"
                "Is a tornado heading towards Toyo?\n\n"
                "Are you sure that we still need to remain open at this point?",
                parent=parent):
            return

        # "Very well then..." popup — same shape/feel as a messagebox but
        # with NO OK button. It auto-dismisses after ~1.5 seconds and then
        # chains into the loading screen. Modal, so only one popup is on
        # screen at a time (the askyesno already closed when Yes was hit).
        notice = tk.Toplevel(parent)
        notice.title("...")
        notice.resizable(False, False)
        notice.transient(parent)
        notice.grab_set()
        notice.protocol("WM_DELETE_WINDOW", lambda: None)
        # Center over the parent
        try:
            parent.update_idletasks()
            px = parent.winfo_rootx() + parent.winfo_width() // 2
            py = parent.winfo_rooty() + parent.winfo_height() // 2
            notice.geometry(f"260x110+{px - 130}+{py - 55}")
        except tk.TclError:
            notice.geometry("260x110")
        ttk.Label(notice, text="Very well then...",
                  font=("Helvetica", 12)).pack(expand=True, padx=20, pady=20)

        def _continue():
            try:
                notice.destroy()
            except tk.TclError:
                pass
            self._open_abyss_loader(parent)

        notice.after(1500, _continue)
        return

    def _open_abyss_loader(self, parent):
        # Black 8-bit-style loading screen with the animated abyss text
        loading = tk.Toplevel(parent)
        loading.title("")
        loading.geometry("560x320")
        loading.configure(bg="black")
        loading.resizable(False, False)
        loading.transient(parent)
        loading.grab_set()
        loading.overrideredirect(False)

        label = tk.Label(loading, text="Entering the Abyss...",
                         font=("Courier", 22, "bold"),
                         fg="white", bg="black")
        label.place(relx=0.5, rely=0.5, anchor="center")

        # Trailing-ellipsis animation: alternates between 3 and 2 dots
        # to match the "... .. ... .. ..." rhythm.
        frames = ["...", " ..", "...", "..", "..."]
        ABYSS_DURATION_MS = 5200   # original 3200 + 2000 longer
        INTERVAL_MS = 380
        state = {"i": 0, "elapsed": 0}

        def tick():
            if state["elapsed"] >= ABYSS_DURATION_MS:
                try:
                    loading.destroy()
                except tk.TclError:
                    pass
                # Close the staff dialog (parent) so it doesn't lurk behind
                # the game window once the easter egg launches.
                try:
                    if parent is not self.root:
                        parent.destroy()
                except tk.TclError:
                    pass
                # Pass DataManager so the game's win-screen recruit popup
                # can add Shadow Realm staff to the roster, plus a callback
                # so the Staff treeview refreshes immediately when a new
                # recruit is added.
                _ExplorerGame(self.root, data=self.data,
                              on_recruit=lambda s: self._refresh_staff_tree())
                return
            f = frames[state["i"] % len(frames)]
            label.config(text=f"Entering the Abyss{f}")
            state["i"] += 1
            state["elapsed"] += INTERVAL_MS
            loading.after(INTERVAL_MS, tick)

        tick()

    def _staff_dialog(self, existing):
        dlg = tk.Toplevel(self.root)
        dlg.title("Edit Staff" if existing else "Add Staff")
        dlg.geometry("520x680")
        dlg.transient(self.root)
        # Defer the grab until Tk has actually mapped the window — calling
        # grab_set() on an unmapped Toplevel raises "window not viewable"
        # and unwinds the rest of this function, leaving a blank dialog.
        dlg.after(50, lambda: dlg.grab_set() if dlg.winfo_exists() else None)

        row = 0
        ttk.Label(dlg, text="Name:").grid(row=row, column=0, padx=10, pady=5, sticky="w")
        name_var = tk.StringVar(value=existing["name"] if existing else "")
        ttk.Entry(dlg, textvariable=name_var, width=25).grid(row=row, column=1, padx=10, pady=5)

        row += 1
        ttk.Label(dlg, text="Roles:").grid(row=row, column=0, padx=10, pady=5, sticky="w")
        roles_frame = ttk.Frame(dlg)
        roles_frame.grid(row=row, column=1, padx=10, pady=5, sticky="w")
        server_var = tk.BooleanVar(value="server" in existing["roles"] if existing else False)
        host_var = tk.BooleanVar(value="host" in existing["roles"] if existing else False)
        manager_var = tk.BooleanVar(value="manager" in existing["roles"] if existing else False)
        ttk.Checkbutton(roles_frame, text="Server", variable=server_var).pack(side=tk.LEFT)
        ttk.Checkbutton(roles_frame, text="Host", variable=host_var).pack(side=tk.LEFT)
        ttk.Checkbutton(roles_frame, text="Manager", variable=manager_var).pack(side=tk.LEFT)

        row += 1
        ttk.Label(dlg, text="Seniority (1-10):").grid(row=row, column=0, padx=10, pady=5, sticky="w")
        sen_var = tk.IntVar(value=existing["seniority"] if existing else 5)
        ttk.Spinbox(dlg, from_=1, to=10, textvariable=sen_var, width=5).grid(row=row, column=1, padx=10, pady=5, sticky="w")

        row += 1
        ttk.Label(dlg, text="Role Preference:").grid(row=row, column=0, padx=10, pady=5, sticky="w")
        pref_var = tk.StringVar(value=existing.get("role_preference", "") or "none" if existing else "none")
        pref_combo = ttk.Combobox(dlg, textvariable=pref_var,
                                  values=["none", "server", "host", "manager"], state="readonly", width=10)
        pref_combo.grid(row=row, column=1, padx=10, pady=5, sticky="w")

        row += 1
        ttk.Label(dlg, text="Flags:").grid(row=row, column=0, padx=10, pady=5, sticky="nw")
        help_btn = tk.Button(
            dlg, text="❓  Staff Help (flags + day offs)",
            font=("Helvetica", 10, "bold"),
            bg="#FFD54F", fg="black", activebackground="#FFC107",
            relief="raised", bd=2, cursor="hand2",
            command=lambda: self._show_flags_help(dlg))
        help_btn.grid(row=row, column=1, padx=10, pady=(5, 2),
                      sticky="w", ipadx=8, ipady=4)

        row += 1
        flags_frame = ttk.Frame(dlg)
        flags_frame.grid(row=row, column=1, padx=10, pady=(2, 5), sticky="w")
        existing_flags = existing.get("flags", []) if existing else []
        flag_vars = {}
        for flag in ["fixed_schedule", "always_hibachi", "no_closing",
                      "fill_in", "emergency_only", "seniority_priority"]:
            var = tk.BooleanVar(value=flag in existing_flags)
            ttk.Checkbutton(flags_frame, text=flag, variable=var).pack(anchor="w")
            flag_vars[flag] = var

        row += 1
        # Easter egg: looks like a real "in case of crisis" override but
        # actually triggers a confirmation flow into a tiny mini-game.
        tk.Button(dlg, text="⚠  ABSOLUTE EMERGENCY  ⚠",
                  font=("Helvetica", 10, "bold"),
                  bg="#C62828", fg="white",
                  activebackground="#8E0000", activeforeground="white",
                  relief="raised", bd=2, cursor="hand2",
                  command=lambda: self._absolute_emergency(dlg)).grid(
            row=row, column=1, padx=10, pady=(8, 6), sticky="w",
            ipadx=8, ipady=3)

        row += 1
        ttk.Label(dlg, text="Fixed Schedule:").grid(row=row, column=0, padx=10, pady=5, sticky="w")
        fixed_var = tk.BooleanVar(value=existing.get("fixed_schedule", False) if existing else False)
        ttk.Checkbutton(dlg, variable=fixed_var).grid(row=row, column=1, padx=10, pady=5, sticky="w")

        row += 1
        ttk.Label(dlg, text="Active:").grid(row=row, column=0, padx=10, pady=5, sticky="w")
        active_var = tk.BooleanVar(value=existing["active"] if existing else True)
        ttk.Checkbutton(dlg, variable=active_var).grid(row=row, column=1, padx=10, pady=5, sticky="w")

        # Day offs — recurring days this staff member doesn't work. Used as
        # a default when generating availability and as a permanent off-day
        # for fixed-schedule staff.
        row += 1
        ttk.Label(dlg, text="Day Offs:").grid(
            row=row, column=0, padx=10, pady=(8, 2), sticky="nw")
        offs_frame = ttk.Frame(dlg)
        offs_frame.grid(row=row, column=1, padx=10, pady=(8, 2), sticky="w")
        existing_offs = set(existing.get("default_off", []) if existing else [])
        day_off_vars = {}
        for di, day in enumerate(DAYS):
            v = tk.BooleanVar(value=(day in existing_offs))
            ttk.Checkbutton(offs_frame, text=day, variable=v).grid(
                row=di // 4, column=di % 4, padx=4, pady=2, sticky="w")
            day_off_vars[day] = v

        def save():
            name = name_var.get().strip()
            if not name:
                messagebox.showerror("Error", "Name is required.", parent=dlg)
                return
            roles = []
            if server_var.get():
                roles.append("server")
            if host_var.get():
                roles.append("host")
            if manager_var.get():
                roles.append("manager")
            if not roles:
                messagebox.showerror("Error", "Select at least one role.", parent=dlg)
                return

            flags = [f for f, v in flag_vars.items() if v.get()]
            # Preserve flags managed outside this dialog (e.g. the exclusive
            # ABSOLUTE EMERGENCY flag from Shadow Realm recruits) so they
            # don't get wiped on edit.
            if existing:
                known_dialog_flags = set(flag_vars.keys())
                for f in existing.get("flags", []):
                    if f not in known_dialog_flags and f not in flags:
                        flags.append(f)
            pref = pref_var.get()
            if pref == "none":
                pref = None

            day_offs = [d for d in DAYS if day_off_vars[d].get()]

            new_staff = {
                "name": name,
                "roles": roles,
                "seniority": sen_var.get(),
                "fixed_schedule": fixed_var.get(),
                "active": active_var.get(),
                "flags": flags,
                "role_preference": pref,
                "default_off": day_offs,
            }
            # Preserve default_availability (and any other unknown fields)
            # from the existing record so they don't get wiped on edit.
            if existing:
                for k, v in existing.items():
                    if k not in new_staff:
                        new_staff[k] = v

            if existing:
                # Update in place
                for i, s in enumerate(self.data.staff):
                    if s["name"] == existing["name"]:
                        self.data.staff[i] = new_staff
                        break
            else:
                # Check duplicate
                if self.data.get_staff_by_name(name):
                    messagebox.showerror("Error", f"{name} already exists.", parent=dlg)
                    return
                self.data.staff.append(new_staff)

            self.data.save_staff()
            self._refresh_staff_tree()
            dlg.destroy()

        row += 1
        ttk.Button(dlg, text="Save", command=save).grid(row=row, column=0, columnspan=2, pady=15)

    # ---- TAB 2: SHIFT CONFIGURATION ----
    def _build_config_tab(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="  Configuration  ")

        # Staffing grid — days on the X axis, role/shift categories on the Y axis.
        # Servers + hibachi rows go on top (frequently edited); hosts go at the
        # bottom (rarely edited). Lunch rows are orange, dinner rows are blue —
        # matching the Availability tab color coding. The grid expands with the
        # window: starts at a comfortable midsize and uses extra space when the
        # window is enlarged.
        style = ttk.Style()
        style.configure("Config.TLabelframe.Label",
                        font=("Helvetica", 13, "bold"))
        style.configure("Config.TButton", font=("Helvetica", 12))

        wrapper = ttk.Frame(frame)
        wrapper.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)
        grid_frame = ttk.LabelFrame(wrapper, text="Staffing Numbers Per Day",
                                    padding=12, style="Config.TLabelframe")
        grid_frame.pack(fill=tk.BOTH, expand=True)

        LUNCH_COLOR = "darkorange"
        DINNER_COLOR = "darkblue"

        # (key, label, color) — order on screen, top to bottom
        rows = [
            ("lunch_servers",  "Lunch Servers",  LUNCH_COLOR),
            ("hibachi_lunch",  "Hibachi Lunch",  LUNCH_COLOR),
            ("dinner_servers", "Dinner Servers", DINNER_COLOR),
            ("hibachi_dinner", "Hibachi Dinner", DINNER_COLOR),
            # --- separator row goes here ---
            ("lunch_hosts",    "Lunch Hosts",    LUNCH_COLOR),
            ("dinner_hosts",   "Dinner Hosts",   DINNER_COLOR),
        ]
        SEP_AFTER_INDEX = 4  # insert visual separator after the 4th data row

        header_font = ("Helvetica", 14, "bold")
        row_label_font = ("Helvetica", 13, "bold")
        spinbox_font = ("Helvetica", 15)
        cell_padx = 12
        cell_pady = 9

        # Top-left empty corner
        ttk.Label(grid_frame, text="", width=16).grid(
            row=0, column=0, padx=cell_padx, pady=cell_pady, sticky="nsew")
        # Day column headers across the top (X axis)
        for di, day in enumerate(DAYS):
            ttk.Label(grid_frame, text=day, font=header_font,
                      anchor="center").grid(
                row=0, column=di + 1, padx=cell_padx, pady=cell_pady, sticky="nsew")

        # Equal-width day columns that grow with the window
        for di in range(len(DAYS)):
            grid_frame.columnconfigure(di + 1, weight=1, uniform="day", minsize=72)

        self.config_vars = {day: {} for day in DAYS}

        # Live-sync helper: any spinbox change immediately writes through to
        # self.data.config so the schedule generator picks up new values
        # without needing the user to click Save Configuration first.
        def make_live_sync(d, k, v):
            def sync(*_args):
                try:
                    self.data.config["staffing"][d][k] = int(v.get())
                except (tk.TclError, ValueError):
                    pass  # transient bad value while user is typing
            return sync

        # Place rows (with a separator inserted between servers/hibachi and hosts)
        grid_row = 1
        data_row_indices = []
        for ri, (key, label, color) in enumerate(rows):
            if ri == SEP_AFTER_INDEX:
                ttk.Separator(grid_frame, orient="horizontal").grid(
                    row=grid_row, column=0, columnspan=len(DAYS) + 1,
                    sticky="ew", pady=6)
                grid_row += 1

            tk.Label(grid_frame, text=label, font=row_label_font,
                     fg=color, anchor="w").grid(
                row=grid_row, column=0, padx=cell_padx, pady=cell_pady, sticky="w")
            for di, day in enumerate(DAYS):
                val = self.data.config["staffing"][day].get(key, 0)
                var = tk.IntVar(value=val)
                self.config_vars[day][key] = var
                var.trace_add("write", make_live_sync(day, key, var))
                sb = tk.Spinbox(grid_frame, from_=0, to=15, textvariable=var,
                                width=5, justify="center", font=spinbox_font)
                sb.grid(row=grid_row, column=di + 1,
                        padx=cell_padx, pady=cell_pady)
            data_row_indices.append(grid_row)
            grid_row += 1

        # Make data rows share extra vertical space when the window grows
        for r in data_row_indices:
            grid_frame.rowconfigure(r, weight=1)

        # Save / Revert buttons
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill=tk.X, padx=15, pady=10)
        ttk.Button(btn_frame, text="Save Configuration",
                   command=self._save_config,
                   style="Config.TButton").pack(side=tk.RIGHT, padx=4, ipadx=8, ipady=4)
        ttk.Button(btn_frame, text="Revert",
                   command=self._revert_config,
                   style="Config.TButton").pack(side=tk.RIGHT, padx=4, ipadx=8, ipady=4)
        ttk.Button(btn_frame, text="Copy Weekday to All",
                   command=self._copy_weekday,
                   style="Config.TButton").pack(side=tk.LEFT, padx=5, ipadx=8, ipady=4)

    def _save_config(self):
        # Spinbox values live-sync into self.data.config via trace_add, so the
        # in-memory state is always current. Save Configuration just persists
        # that state to disk for future sessions.
        self.data.save_config()
        messagebox.showinfo("Saved", "Configuration saved!")

    def _revert_config(self):
        if not messagebox.askyesno(
                "Revert Configuration",
                "Are you sure you want to revert the configuration to the last "
                "saved version? Any unsaved changes will be lost."):
            return
        # Reload from disk and push the values back into the spinbox vars.
        # The trace on each var will update self.data.config to the same value.
        self.data.config = self.data._load_config()
        for day in DAYS:
            for key, var in self.config_vars[day].items():
                var.set(self.data.config["staffing"][day].get(key, 0))

    def _copy_weekday(self):
        """Copy Monday values to Tue-Wed only. Thursday is treated as part of
        the busy weekend block (Thu-Sun) and is left untouched."""
        keys = ["lunch_servers", "lunch_hosts", "hibachi_lunch",
                "dinner_servers", "dinner_hosts", "hibachi_dinner"]
        for day in ["TUE", "WED"]:
            for key in keys:
                self.config_vars[day][key].set(self.config_vars["MON"][key].get())

    # ---- TAB 3: AVAILABILITY ----
    def _build_availability_tab(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="  Availability  ")

        # Toolbar
        toolbar = ttk.Frame(frame)
        toolbar.pack(fill=tk.X, padx=8, pady=8)
        ttk.Button(toolbar, text="Set Default", command=self._set_default_availability,
                   style="Big.TButton").pack(side=tk.LEFT, padx=3, ipadx=6, ipady=3)
        ttk.Button(toolbar, text="Save Availability", command=self._save_availability,
                   style="Big.TButton").pack(side=tk.RIGHT, padx=3, ipadx=6, ipady=3)
        ttk.Button(toolbar, text="Revert", command=self._revert_availability,
                   style="Big.TButton").pack(side=tk.RIGHT, padx=3, ipadx=6, ipady=3)

        # Scrollable grid
        canvas = tk.Canvas(frame)
        v_scroll = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=canvas.yview)
        self.avail_frame = ttk.Frame(canvas)

        self.avail_frame.bind("<Configure>",
                              lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.avail_frame, anchor="nw")
        canvas.configure(yscrollcommand=v_scroll.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        v_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # Bind mousewheel
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        canvas.bind_all("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))
        canvas.bind_all("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))

        self._build_availability_grid()

    def _build_availability_grid(self):
        for widget in self.avail_frame.winfo_children():
            widget.destroy()

        # Internal avail_vars still tracks name -> day -> StringVar("off"/"morning"/"night"/"both")
        # for compatibility with save/load/photo import
        self.avail_vars = {}
        active = [s for s in self.data.staff if s["active"]]

        # Initialize avail_vars for all active non-manager staff
        for staff in active:
            if "manager" in staff["roles"]:
                continue
            name = staff["name"]
            if name not in self.avail_vars:
                self.avail_vars[name] = {}
                for day in DAYS:
                    saved_val = self.data.availability.get(name, {}).get(day, "off")
                    if staff.get("fixed_schedule") and saved_val == "off":
                        default_avail = staff.get("default_availability", {})
                        if default_avail and day in default_avail:
                            saved_val = default_avail[day]
                        else:
                            default_off = staff.get("default_off", [])
                            saved_val = "off" if day in default_off else "both"
                    var = tk.StringVar(value=saved_val)
                    self.avail_vars[name][day] = var
                    # Live-sync this StringVar into self.data.availability so the
                    # generator picks up cell edits without needing Save Availability.
                    def make_live_sync(n=name, d=day, v=var):
                        def sync(*_args):
                            self.data.availability.setdefault(n, {})[d] = v.get()
                        return sync
                    var.trace_add("write", make_live_sync())
                    # Also seed the value into self.data.availability now in case
                    # this is the first time the staff member appears (no entry yet).
                    self.data.availability.setdefault(name, {})[day] = saved_val

        # Build name lists for dropdowns (include emergency/fill-in for manual assignment)
        servers = [s for s in active if "server" in s["roles"] and "manager" not in s["roles"]]
        hosts = [s for s in active if "host" in s["roles"] and "manager" not in s["roles"]]
        servers.sort(key=lambda s: s["name"])
        hosts.sort(key=lambda s: s["name"])

        def make_name_list(staff_list):
            """Build dropdown options: plain names."""
            names = [""]
            for s in staff_list:
                names.append(s["name"])
            return names

        server_names = make_name_list(servers)
        host_names = make_name_list(hosts)

        # Cell widgets and name lists: (role, shift, day) -> cell widget / name list
        self._cell_widgets = {}
        self._slot_names = {}

        # Pre-populate slot_names from avail_vars
        def find_display_name(real_name, name_list):
            if real_name in name_list:
                return real_name
            return real_name

        for role_key, staff_list, name_list in [
            ("server", servers, server_names),
            ("host", hosts, host_names),
        ]:
            for staff in staff_list:
                name = staff["name"]
                # Dual-role staff: only show in their preferred role section
                is_dual = "server" in staff["roles"] and "host" in staff["roles"]
                if is_dual and staff.get("role_preference") and staff["role_preference"] != role_key:
                    continue
                if name not in self.avail_vars:
                    continue
                display = find_display_name(name, name_list)
                for day in DAYS:
                    val = self.avail_vars[name][day].get()
                    if val in ("morning", "both"):
                        key = (role_key, "morning", day)
                        if key not in self._slot_names:
                            self._slot_names[key] = []
                        if display not in self._slot_names[key]:
                            self._slot_names[key].append(display)
                    if val in ("night", "both"):
                        key = (role_key, "night", day)
                        if key not in self._slot_names:
                            self._slot_names[key] = []
                        if display not in self._slot_names[key]:
                            self._slot_names[key].append(display)

        def make_staff_cell(parent, names_list, full_name_list, role_key, shift, day, row, col):
            """Cell with selected names (with X to remove) and a dropdown to add."""
            cell = tk.Frame(parent, relief="groove", bd=1, bg="white", width=160)
            cell.grid(row=row, column=col, padx=3, pady=3, sticky="nsew")
            cell._names = names_list  # list of currently selected names
            cell._full_list = full_name_list
            cell._role_key = role_key
            cell._shift = shift
            cell._day = day

            def extract_name(display_val):
                if not display_val:
                    return ""
                return display_val

            def sync_to_avail():
                """Sync this cell's selections back to avail_vars."""
                # This cell's names contribute to availability
                for display in cell._names:
                    real_name = extract_name(display)
                    if real_name and real_name in self.avail_vars:
                        current = self.avail_vars[real_name][day].get()
                        if shift == "morning":
                            if current in ("night", "both"):
                                self.avail_vars[real_name][day].set("both")
                            else:
                                self.avail_vars[real_name][day].set("morning")
                        else:
                            if current in ("morning", "both"):
                                self.avail_vars[real_name][day].set("both")
                            else:
                                self.avail_vars[real_name][day].set("night")

            def refresh_cell():
                # Clear cell contents
                for w in cell.winfo_children():
                    w.destroy()

                # Show selected names with X buttons
                for i, display_name in enumerate(cell._names):
                    name_frame = tk.Frame(cell, bg="white")
                    name_frame.pack(fill=tk.X, padx=3, pady=2)

                    tk.Label(name_frame, text=display_name, font=("Helvetica", 11),
                             bg="white", anchor="w").pack(side=tk.LEFT, fill=tk.X, expand=True)

                    def make_remove(idx):
                        def remove(event=None):
                            removed = cell._names.pop(idx)
                            real_name = extract_name(removed)
                            # Clear this shift from avail_vars, recalc from all cells
                            if real_name in self.avail_vars:
                                self.avail_vars[real_name][day].set("off")
                                # Re-sync all cells for this name/day
                                for key, c in self._cell_widgets.items():
                                    rk, sh, d = key
                                    if d == day:
                                        for dn in c._names:
                                            rn = extract_name(dn)
                                            if rn == real_name:
                                                cur = self.avail_vars[rn][d].get()
                                                if sh == "morning":
                                                    self.avail_vars[rn][d].set(
                                                        "both" if cur == "night" else "morning")
                                                else:
                                                    self.avail_vars[rn][d].set(
                                                        "both" if cur == "morning" else "night")
                            refresh_cell()
                        return remove

                    x_btn = tk.Label(name_frame, text="x", font=("Helvetica", 11, "bold"),
                                     fg="red", bg="white", cursor="hand2")
                    x_btn.pack(side=tk.RIGHT, padx=3)
                    x_btn.bind("<Button-1>", make_remove(i))

                # Add dropdown button
                add_frame = tk.Frame(cell, bg="white")
                add_frame.pack(fill=tk.X, padx=3, pady=(3, 1))

                add_btn = tk.Label(add_frame, text="+ Add", font=("Helvetica", 10),
                                   fg="gray", bg="#f0f0f0", relief="raised", bd=1,
                                   cursor="hand2")
                add_btn.pack(fill=tk.X, ipady=2)

                def open_popup(event=None):
                    # Toggle: clicking the same +Add closes its open popup.
                    if self._active_add_popup is not None:
                        prev_btn = self._active_add_btn
                        try:
                            self._active_add_popup.destroy()
                        except tk.TclError:
                            pass
                        self._active_add_popup = None
                        self._active_add_btn = None
                        if prev_btn is add_btn:
                            return  # toggled off

                    popup = tk.Toplevel(parent)
                    popup.overrideredirect(True)
                    popup.attributes("-topmost", True)

                    x = add_btn.winfo_rootx()
                    y = add_btn.winfo_rooty() + add_btn.winfo_height()
                    popup.geometry(f"+{x}+{y}")

                    self._active_add_popup = popup
                    self._active_add_btn = add_btn

                    lb_frame = tk.Frame(popup)
                    lb_frame.pack(fill=tk.BOTH, expand=True)

                    scrollbar = tk.Scrollbar(lb_frame)
                    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

                    # Filter out already-selected names
                    selected_real = {extract_name(n) for n in cell._names}
                    available = [n for n in cell._full_list
                                 if n and extract_name(n) not in selected_real]

                    lb = tk.Listbox(lb_frame, width=18, height=min(12, max(len(available), 1)),
                                    font=("Helvetica", 9), selectmode=tk.SINGLE,
                                    yscrollcommand=scrollbar.set, activestyle="dotbox")
                    lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
                    scrollbar.config(command=lb.yview)

                    for name in available:
                        lb.insert(tk.END, name)

                    def on_select(event=None):
                        sel = lb.curselection()
                        if sel:
                            chosen = lb.get(sel[0])
                            if chosen:
                                cell._names.append(chosen)
                                sync_to_avail()
                                refresh_cell()
                        popup.destroy()

                    def on_key(event):
                        ch = event.char
                        if not ch or not ch.isalpha():
                            if event.keysym == 'Return':
                                on_select()
                                return "break"
                            if event.keysym == 'Escape':
                                popup.destroy()
                                return "break"
                            return
                        ch = ch.lower()
                        for i in range(lb.size()):
                            item = lb.get(i)
                            if item and item.lower().startswith(ch):
                                lb.see(i)
                                lb.selection_clear(0, tk.END)
                                lb.selection_set(i)
                                lb.activate(i)
                                return "break"
                        return "break"

                    def on_arrow(event):
                        idx = lb.index(tk.ACTIVE)
                        if event.keysym == 'Down' and idx < lb.size() - 1:
                            idx += 1
                        elif event.keysym == 'Up' and idx > 0:
                            idx -= 1
                        lb.selection_clear(0, tk.END)
                        lb.selection_set(idx)
                        lb.activate(idx)
                        lb.see(idx)
                        return "break"

                    lb.bind("<ButtonRelease-1>", on_select)
                    lb.bind("<Key>", on_key)
                    lb.bind("<Down>", on_arrow)
                    lb.bind("<Up>", on_arrow)
                    lb.bind("<Escape>", lambda e: popup.destroy())
                    popup.update_idletasks()
                    lb.focus_force()

                    # Close popup when clicking anywhere outside it.
                    # If the click lands on the active +Add button we let its
                    # own handler run so it can toggle this popup off cleanly.
                    def _make_close_handler(p, toplevel):
                        def handler(event):
                            try:
                                if event.widget is self._active_add_btn:
                                    return
                                p_str = str(p)
                                w_str = str(event.widget)
                                if w_str != p_str and not w_str.startswith(p_str + "."):
                                    p.destroy()
                            except tk.TclError:
                                pass
                        def cleanup(event):
                            if event.widget == p:
                                try:
                                    toplevel.unbind("<Button-1>")
                                except tk.TclError:
                                    pass
                                if self._active_add_popup is p:
                                    self._active_add_popup = None
                                    self._active_add_btn = None
                                # The listbox stole keyboard focus via
                                # focus_force on an overrideredirect popup.
                                # Without this restore the main window
                                # ignores clicks until the user clicks
                                # outside the app.
                                try:
                                    toplevel.focus_force()
                                except tk.TclError:
                                    pass
                        return handler, cleanup
                    _toplevel = parent.winfo_toplevel()
                    _click_handler, _destroy_handler = _make_close_handler(popup, _toplevel)
                    _toplevel.bind("<Button-1>", _click_handler, add="+")
                    popup.bind("<Destroy>", _destroy_handler)

                add_btn.bind("<Button-1>", open_popup)

                # Typing textbox: type space-separated tokens like "ka kam sa".
                # Each token resolves to the shortest name in full_list that
                # starts with it (case-insensitive). Hit Enter to add them all.
                # A live preview label shows what the tokens will resolve to.
                type_entry = tk.Entry(cell, font=("Helvetica", 11), bd=1, relief="sunken")
                type_entry.pack(fill=tk.X, padx=3, pady=(2, 0), ipady=2)
                cell._type_entry = type_entry

                preview_label = tk.Label(cell, text="", font=("Helvetica", 9, "italic"),
                                         fg="#0066CC", bg="white", anchor="w")
                preview_label.pack(fill=tk.X, padx=3, pady=(0, 3))

                def resolve_token(tok):
                    tok_low = tok.lower()
                    matches = [n for n in cell._full_list
                               if n and n.lower().startswith(tok_low)]
                    if not matches:
                        return None
                    # Shortest first; alphabetical as deterministic tiebreaker
                    matches.sort(key=lambda n: (len(n), n.lower()))
                    return matches[0]

                def update_preview(event=None):
                    raw = type_entry.get().strip()
                    if not raw:
                        preview_label.config(text="", fg="#0066CC")
                        return
                    tokens = raw.split()
                    parts = []
                    any_miss = False
                    for tok in tokens:
                        m = resolve_token(tok)
                        if m:
                            parts.append(m)
                        else:
                            parts.append(f"?{tok}")
                            any_miss = True
                    preview_label.config(
                        text="→ " + ", ".join(parts),
                        fg="#CC0000" if any_miss else "#0066CC")

                def on_type_enter(event=None):
                    raw = type_entry.get().strip()
                    if not raw:
                        return "break"
                    tokens = raw.split()
                    selected_real = {extract_name(n) for n in cell._names}
                    added_any = False
                    for tok in tokens:
                        match = resolve_token(tok)
                        if not match or match in selected_real:
                            continue
                        cell._names.append(match)
                        selected_real.add(match)
                        added_any = True
                    type_entry.delete(0, tk.END)
                    preview_label.config(text="", fg="#0066CC")
                    if added_any:
                        sync_to_avail()
                        refresh_cell()
                        # refresh_cell rebuilds widgets — focus the NEW entry
                        # so the user can immediately type the next name.
                        cell._type_entry.focus_set()
                    return "break"

                type_entry.bind("<KeyRelease>", update_preview)
                type_entry.bind("<Return>", on_type_enter)

            refresh_cell()
            return cell

        def add_shift_section(row, title, name_list, role_key, shift, color):
            """Add a single shift-role section (e.g. Host Morning, Lunch Server)."""
            # Section header
            ttk.Label(self.avail_frame, text=title,
                      font=("Helvetica", 14, "bold"), foreground="navy").grid(
                row=row, column=0, columnspan=8, padx=8, pady=(12, 4), sticky="w")
            row += 1

            # Day column headers
            ttk.Label(self.avail_frame, text="", width=10).grid(row=row, column=0)
            for di, day in enumerate(DAYS):
                ttk.Label(self.avail_frame, text=day, font=("Helvetica", 12, "bold"),
                          width=16).grid(row=row, column=di + 1, padx=2, pady=4)
            row += 1

            # Staff cells
            ttk.Label(self.avail_frame, text=shift.capitalize(),
                      font=("Helvetica", 12, "bold"), foreground=color).grid(
                row=row, column=0, padx=8, pady=(6, 2), sticky="nw")
            for di, day in enumerate(DAYS):
                key = (role_key, shift, day)
                names = self._slot_names.get(key, [])
                cell = make_staff_cell(self.avail_frame, names, name_list,
                                       role_key, shift, day, row, di + 1)
                self._cell_widgets[key] = cell
                self._slot_names[key] = names
            row += 1

            return row

        row = 0

        # Layout matches the sign-up sheet order:
        # 1. Host Morning
        row = add_shift_section(row, "HOST — MORNING", host_names, "host", "morning", "darkorange")

        ttk.Separator(self.avail_frame, orient="horizontal").grid(
            row=row, column=0, columnspan=8, sticky="ew", pady=8)
        row += 1

        # 2. Lunch Servers
        row = add_shift_section(row, "SERVER — LUNCH", server_names, "server", "morning", "darkorange")

        ttk.Separator(self.avail_frame, orient="horizontal").grid(
            row=row, column=0, columnspan=8, sticky="ew", pady=8)
        row += 1

        # 3. Host Night
        row = add_shift_section(row, "HOST — DINNER", host_names, "host", "night", "darkblue")

        ttk.Separator(self.avail_frame, orient="horizontal").grid(
            row=row, column=0, columnspan=8, sticky="ew", pady=8)
        row += 1

        # 4. Dinner Servers
        row = add_shift_section(row, "SERVER — DINNER", server_names, "server", "night", "darkblue")

    def _set_default_availability(self):
        """Reset everyone to off, then restore the fulltimers (fixed-schedule
        staff) to their default availability. Used as a clean starting point
        for filling in the rest of the week."""
        if not messagebox.askyesno(
                "Set Default",
                "Are you sure you want to clear everyone and restore the "
                "fulltimers' default schedules? Any unsaved availability "
                "edits will be lost."):
            return
        for staff in self.data.staff:
            if not staff["active"] or "manager" in staff["roles"]:
                continue
            name = staff["name"]
            if staff.get("fixed_schedule"):
                default_avail = staff.get("default_availability", {})
                for day in DAYS:
                    if default_avail and day in default_avail:
                        self.data.availability.setdefault(name, {})[day] = default_avail[day]
                    else:
                        default_off = staff.get("default_off", [])
                        self.data.availability.setdefault(name, {})[day] = "off" if day in default_off else "both"
            else:
                for day in DAYS:
                    self.data.availability.setdefault(name, {})[day] = "off"
        self._build_availability_grid()

    def _save_availability(self):
        # avail_vars live-sync into self.data.availability via trace_add as soon
        # as cells are edited, so the in-memory state is always current. Save
        # Availability just persists that state to disk for future sessions.
        # We still do a defensive cell→avail_vars rebuild here in case anything
        # got out of sync (the trace then propagates it to self.data.availability).
        if hasattr(self, '_cell_widgets'):
            def extract_name(display_val):
                if not display_val:
                    return ""
                return display_val

            # Reset all non-fixed to off, then rebuild from current cell contents
            for name, days in self.avail_vars.items():
                staff = self.data.get_staff_by_name(name)
                if staff and staff.get("fixed_schedule"):
                    continue
                for day in DAYS:
                    days[day].set("off")

            for (role_key, shift, day), cell in self._cell_widgets.items():
                for display in cell._names:
                    real_name = extract_name(display)
                    if real_name and real_name in self.avail_vars:
                        current = self.avail_vars[real_name][day].get()
                        if shift == "morning":
                            if current in ("night", "both"):
                                self.avail_vars[real_name][day].set("both")
                            else:
                                self.avail_vars[real_name][day].set("morning")
                        else:
                            if current in ("morning", "both"):
                                self.avail_vars[real_name][day].set("both")
                            else:
                                self.avail_vars[real_name][day].set("night")

        self.data.save_availability()
        messagebox.showinfo("Saved", "Availability saved!")

    def _revert_availability(self):
        if not messagebox.askyesno(
                "Revert Availability",
                "Are you sure you want to revert availability to the last "
                "saved version? Any unsaved changes will be lost."):
            return
        # Reload from disk and rebuild the grid (which reads self.data.availability)
        self.data.availability = self.data._load_availability()
        self._build_availability_grid()

    # ---- TAB 4: SCHEDULE GENERATION ----
    def _build_schedule_tab(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="  Schedule  ")

        # Toolbar
        toolbar = ttk.Frame(frame)
        toolbar.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(toolbar, text="Week Starting (Mon):").pack(side=tk.LEFT, padx=5)
        self.week_start_var = tk.StringVar(value=self.data.config.get("week_start", ""))
        ttk.Entry(toolbar, textvariable=self.week_start_var, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Label(toolbar, text="(YYYY-MM-DD)").pack(side=tk.LEFT, padx=2)

        ttk.Button(toolbar, text="Generate Schedule",
                  command=self._generate_schedule).pack(side=tk.LEFT, padx=15)

        # Warning area
        self.warning_var = tk.StringVar(value="")
        self.warning_label = ttk.Label(frame, textvariable=self.warning_var,
                                       foreground="red", wraplength=1100)
        self.warning_label.pack(fill=tk.X, padx=10)

        # Schedule display - scrollable canvas with a table
        canvas_frame = ttk.Frame(frame)
        canvas_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.sched_canvas = tk.Canvas(canvas_frame)
        v_scroll = ttk.Scrollbar(canvas_frame, orient=tk.VERTICAL, command=self.sched_canvas.yview)
        h_scroll = ttk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL, command=self.sched_canvas.xview)
        self.sched_inner = ttk.Frame(self.sched_canvas)
        self.sched_inner.bind("<Configure>",
                              lambda e: self.sched_canvas.configure(scrollregion=self.sched_canvas.bbox("all")))
        self.sched_canvas.create_window((0, 0), window=self.sched_inner, anchor="nw")
        self.sched_canvas.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

        self.sched_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        h_scroll.pack(side=tk.BOTTOM, fill=tk.X)

    def _generate_schedule(self):
        # Save availability first
        self._save_availability_silent()
        self._save_config_silent()

        self.current_schedule, self.current_warnings = self.generator.generate()

        if self.current_warnings:
            self.warning_var.set("Warnings: " + " | ".join(self.current_warnings))
        else:
            self.warning_var.set("Schedule generated successfully!")

        self._display_schedule()

    def _save_availability_silent(self):
        if hasattr(self, 'avail_vars'):
            avail = {}
            for name, days in self.avail_vars.items():
                avail[name] = {}
                for day, var in days.items():
                    avail[name][day] = var.get()
            self.data.availability = avail
            self.data.save_availability()

    def _save_config_silent(self):
        if hasattr(self, 'config_vars'):
            keys = ["lunch_servers", "lunch_hosts", "hibachi_lunch",
                    "dinner_servers", "dinner_hosts", "hibachi_dinner"]
            for day in DAYS:
                for key in keys:
                    if key in self.config_vars[day]:
                        self.data.config["staffing"][day][key] = self.config_vars[day][key].get()
            self.data.config["week_start"] = self.week_start_var.get()
            self.data.save_config()

    def _display_schedule(self):
        for widget in self.sched_inner.winfo_children():
            widget.destroy()

        if not self.current_schedule:
            return

        # Header row
        ttk.Label(self.sched_inner, text="", width=18, font=("Helvetica", 9, "bold")).grid(
            row=0, column=0, padx=2, pady=2)
        for di, day in enumerate(DAYS):
            lbl = ttk.Label(self.sched_inner, text=day, width=18,
                           font=("Helvetica", 10, "bold"), anchor="center")
            lbl.grid(row=0, column=di + 1, padx=2, pady=2)

        row = 1
        # Section: LUNCH SERVERS
        ttk.Label(self.sched_inner, text="LUNCH SERVERS",
                 font=("Helvetica", 9, "bold"), foreground="blue").grid(
            row=row, column=0, padx=5, pady=3, sticky="w")
        ttk.Label(self.sched_inner, text=LUNCH_SERVER_TIME,
                 font=("Helvetica", 8)).grid(row=row + 1, column=0, padx=5, sticky="w")
        row += 1

        max_ls = max(len(self.current_schedule[d]["lunch_servers"]) + len(self.current_schedule[d]["hibachi_lunch_servers"]) for d in DAYS)
        for si in range(max(max_ls, 1)):
            for di, day in enumerate(DAYS):
                hibachi_lunch = self.current_schedule[day]["hibachi_lunch_servers"]
                servers = self.current_schedule[day]["lunch_servers"]
                letters = self.current_schedule[day].get("lunch_server_letters", {})
                b_side = self.current_schedule[day].get("b_side_lunch", [])
                hib_set = set(hibachi_lunch)
                b_set = set(b_side)
                # B-side on top, then regular A,B,C..., then hibachi (within letter)
                combined = servers + hibachi_lunch
                combined.sort(key=lambda n: (
                    0 if n in b_set else 1,
                    letters.get(n, "Z"),
                    1 if n in hib_set else 0,
                ))
                mids = self.current_schedule[day].get("mid_servers", [])
                if si < len(combined):
                    name = combined[si]
                    letter = letters.get(name, chr(65 + si))
                    # In-app: midshift shows on EVERY cell the person appears
                    # in for that day (text *, green BG, and the button), unless
                    # the cell is overridden by B-side blue or hibachi orange.
                    is_mid = name in mids
                    mid = "*" if is_mid else ""
                    is_hib = name in hib_set
                    is_b = name in b_set
                    text = f"<{name} {letter}>{mid}" if is_b else f"{name} {letter}{mid}"
                    if is_b:
                        bg = "#ADD8E6"  # light blue for B-side overrides green
                    elif is_hib:
                        bg = "#FFE4B5"
                    elif is_mid:
                        bg = "#C8E6C9"  # green highlight for midshift
                    else:
                        bg = "white"
                    cell = tk.Frame(self.sched_inner, bg=bg, relief="groove", bd=1)
                    cell.grid(row=row + si, column=di + 1, padx=1, pady=1, sticky="nsew")
                    lbl = tk.Label(cell, text=text, width=13,
                                  font=("Helvetica", 9), anchor="center",
                                  bg=bg, cursor="hand2", bd=0)
                    lbl.pack(side="left", fill="both", expand=True)
                    lbl.bind("<Button-1>", lambda e, d=day, k="lunch_servers", i=si: self._edit_cell(d, k, i))
                    btn_bg = "#4A90E2" if is_b else "#DDDDDD"
                    btn_fg = "white" if is_b else "#333333"
                    bbtn = tk.Label(cell, text="B", width=2,
                                    font=("Helvetica", 8, "bold"),
                                    bg=btn_bg, fg=btn_fg, cursor="hand2", bd=1, relief="raised")
                    bbtn.pack(side="right", fill="y")
                    bbtn.bind("<Button-1>", lambda e, d=day, n=name: self._toggle_b_side(d, "lunch", n))
                    mbtn_bg = "#4CAF50" if is_mid else "#DDDDDD"
                    mbtn_fg = "white" if is_mid else "#333333"
                    mbtn = tk.Label(cell, text="*", width=2,
                                    font=("Helvetica", 10, "bold"),
                                    bg=mbtn_bg, fg=mbtn_fg, cursor="hand2", bd=1, relief="raised")
                    mbtn.pack(side="right", fill="y")
                    mbtn.bind("<Button-1>", lambda e, d=day, n=name: self._toggle_mid(d, n))
        row += max(max_ls, 1) + 1

        # Lunch Manager
        ttk.Label(self.sched_inner, text="LUNCH MANAGER",
                 font=("Helvetica", 9, "bold"), foreground="darkgreen").grid(
            row=row, column=0, padx=5, pady=3, sticky="w")
        for di, day in enumerate(DAYS):
            mgr = self.current_schedule[day]["lunch_manager"]
            lbl = tk.Label(self.sched_inner, text=mgr, width=18,
                          font=("Helvetica", 9, "bold"), anchor="center",
                          relief="groove", bd=1)
            lbl.grid(row=row, column=di + 1, padx=1, pady=1)
        row += 1

        # Lunch Hosts
        ttk.Label(self.sched_inner, text="LUNCH HOSTS",
                 font=("Helvetica", 9, "bold"), foreground="teal").grid(
            row=row, column=0, padx=5, pady=3, sticky="w")
        row += 1
        max_lh = max(len(self.current_schedule[d]["lunch_hosts"]) for d in DAYS)
        for hi in range(max(max_lh, 1)):
            for di, day in enumerate(DAYS):
                hosts = self.current_schedule[day]["lunch_hosts"]
                if hi < len(hosts):
                    lbl = tk.Label(self.sched_inner, text=hosts[hi], width=18,
                                  font=("Helvetica", 9), anchor="center",
                                  relief="groove", bd=1, cursor="hand2")
                    lbl.grid(row=row + hi, column=di + 1, padx=1, pady=1)
                    lbl.bind("<Button-1>", lambda e, d=day, k="lunch_hosts", i=hi: self._edit_cell(d, k, i))
        row += max(max_lh, 1) + 1

        # Separator
        ttk.Separator(self.sched_inner, orient="horizontal").grid(
            row=row, column=0, columnspan=8, sticky="ew", pady=5)
        row += 1

        # DINNER SERVERS
        ttk.Label(self.sched_inner, text="DINNER SERVERS",
                 font=("Helvetica", 9, "bold"), foreground="darkred").grid(
            row=row, column=0, padx=5, pady=3, sticky="w")
        ttk.Label(self.sched_inner, text=DINNER_SERVER_TIME,
                 font=("Helvetica", 8)).grid(row=row + 1, column=0, padx=5, sticky="w")
        row += 1

        max_ds = max(
            len(self.current_schedule[d]["dinner_servers"]) +
            len(self.current_schedule[d]["hibachi_servers"])
            for d in DAYS)
        for si in range(max(max_ds, 1)):
            for di, day in enumerate(DAYS):
                hibachi = self.current_schedule[day]["hibachi_servers"]
                servers = self.current_schedule[day]["dinner_servers"]
                letters = self.current_schedule[day].get("dinner_server_letters", {})
                b_side = self.current_schedule[day].get("b_side_dinner", [])
                hib_set = set(hibachi)
                b_set = set(b_side)
                # B-side on top, then regular A,B,C..., then hibachi (within letter)
                combined = servers + hibachi
                combined.sort(key=lambda n: (
                    0 if n in b_set else 1,
                    letters.get(n, "Z"),
                    1 if n in hib_set else 0,
                ))
                mids = self.current_schedule[day].get("mid_servers", [])
                if si < len(combined):
                    name = combined[si]
                    letter = letters.get(name, chr(65 + si))
                    is_mid = name in mids
                    mid = "*" if is_mid else ""
                    is_hib = name in hib_set
                    is_b = name in b_set
                    text = f"<{name} {letter}>{mid}" if is_b else f"{name} {letter}{mid}"
                    if is_b:
                        bg = "#ADD8E6"  # light blue for B-side overrides green
                    elif is_hib:
                        bg = "#FFE4B5"
                    elif is_mid:
                        bg = "#C8E6C9"  # green highlight for midshift
                    else:
                        bg = "white"
                    cell = tk.Frame(self.sched_inner, bg=bg, relief="groove", bd=1)
                    cell.grid(row=row + si, column=di + 1, padx=1, pady=1, sticky="nsew")
                    lbl = tk.Label(cell, text=text, width=13,
                                  font=("Helvetica", 9), anchor="center",
                                  bg=bg, cursor="hand2", bd=0)
                    lbl.pack(side="left", fill="both", expand=True)
                    lbl.bind("<Button-1>", lambda e, d=day, k="dinner_servers", i=si: self._edit_cell(d, k, i))
                    btn_bg = "#4A90E2" if is_b else "#DDDDDD"
                    btn_fg = "white" if is_b else "#333333"
                    bbtn = tk.Label(cell, text="B", width=2,
                                    font=("Helvetica", 8, "bold"),
                                    bg=btn_bg, fg=btn_fg, cursor="hand2", bd=1, relief="raised")
                    bbtn.pack(side="right", fill="y")
                    bbtn.bind("<Button-1>", lambda e, d=day, n=name: self._toggle_b_side(d, "dinner", n))
                    mbtn_bg = "#4CAF50" if is_mid else "#DDDDDD"
                    mbtn_fg = "white" if is_mid else "#333333"
                    mbtn = tk.Label(cell, text="*", width=2,
                                    font=("Helvetica", 10, "bold"),
                                    bg=mbtn_bg, fg=mbtn_fg, cursor="hand2", bd=1, relief="raised")
                    mbtn.pack(side="right", fill="y")
                    mbtn.bind("<Button-1>", lambda e, d=day, n=name: self._toggle_mid(d, n))
        row += max(max_ds, 1) + 1

        # Dinner Hosts
        ttk.Label(self.sched_inner, text="DINNER HOSTS",
                 font=("Helvetica", 9, "bold"), foreground="teal").grid(
            row=row, column=0, padx=5, pady=3, sticky="w")
        row += 1
        max_dh = max(len(self.current_schedule[d]["dinner_hosts"]) for d in DAYS)
        for hi in range(max(max_dh, 1)):
            for di, day in enumerate(DAYS):
                hosts = self.current_schedule[day]["dinner_hosts"]
                if hi < len(hosts):
                    lbl = tk.Label(self.sched_inner, text=hosts[hi], width=18,
                                  font=("Helvetica", 9), anchor="center",
                                  relief="groove", bd=1, cursor="hand2")
                    lbl.grid(row=row + hi, column=di + 1, padx=1, pady=1)
                    lbl.bind("<Button-1>", lambda e, d=day, k="dinner_hosts", i=hi: self._edit_cell(d, k, i))
        row += max(max_dh, 1) + 1

        # Dinner Manager
        ttk.Label(self.sched_inner, text="DINNER MANAGER",
                 font=("Helvetica", 9, "bold"), foreground="darkgreen").grid(
            row=row, column=0, padx=5, pady=3, sticky="w")
        for di, day in enumerate(DAYS):
            mgr = self.current_schedule[day]["dinner_manager"]
            lbl = tk.Label(self.sched_inner, text=mgr, width=18,
                          font=("Helvetica", 9, "bold"), anchor="center",
                          relief="groove", bd=1)
            lbl.grid(row=row, column=di + 1, padx=1, pady=1)
        row += 2

        # Shift count summary
        ttk.Label(self.sched_inner, text="SHIFT TOTALS",
                 font=("Helvetica", 9, "bold")).grid(row=row, column=0, padx=5, pady=5, sticky="w")
        row += 1
        counts = self.generator.shift_counts
        sorted_counts = sorted(counts.items(), key=lambda x: x[1], reverse=True)
        col_offset = 0
        for i, (name, count) in enumerate(sorted_counts):
            r = row + (i % 10)
            c = (i // 10) * 2
            ttk.Label(self.sched_inner, text=f"{name}: {count} shifts",
                     font=("Helvetica", 8)).grid(row=r, column=c, padx=5, sticky="w")

    def _get_available_for_slot(self, day, key):
        """Build list of available staff for a schedule slot, with role labels for dual-role."""
        shift_type = "morning" if "lunch" in key else "night"
        slot_role = "server" if "server" in key else "host"

        avail = self.data.availability
        options = []

        for s in self.data.staff:
            if not s["active"] or "manager" in s["roles"]:
                continue
            name = s["name"]

            # Check availability for this day/shift
            staff_avail = avail.get(name, {}).get(day, "off")
            if staff_avail == "off":
                continue
            if shift_type == "morning" and staff_avail == "night":
                continue
            if shift_type == "night" and staff_avail == "morning":
                continue

            # Check role match and label dual-role staff
            is_dual = "server" in s["roles"] and "host" in s["roles"]
            has_role = slot_role in s["roles"]

            if has_role and is_dual:
                options.append((name, name))
            elif has_role:
                options.append((name, name))

        options.sort(key=lambda x: x[1].lower())
        return options

    def _toggle_mid(self, day, name):
        """Toggle a staff member's midshift status for a day. Pure visual
        marker — no scheduling logic. Shows a green highlight on every cell
        the person appears in that day plus a '*' next to their name on
        their FIRST shift of the day. The Excel export gets the '*' in the
        same place but no fill color."""
        if not self.current_schedule:
            return
        mids = self.current_schedule[day].setdefault("mid_servers", [])
        if name in mids:
            mids.remove(name)
        else:
            mids.append(name)
        self._display_schedule()

    def _toggle_b_side(self, day, shift, name):
        """Toggle a staff member's B-side (sushi side) assignment for a shift.

        shift = 'lunch' or 'dinner'. B-side staff get letters W, X, Y in order
        (max 3 per shift). B-side and hibachi are mutually exclusive — toggling
        someone in pulls them out of hibachi if they were there, and a regular
        server is promoted into the vacated hibachi slot.
        """
        if not self.current_schedule:
            return
        b_key = "b_side_lunch" if shift == "lunch" else "b_side_dinner"
        hib_key = "hibachi_lunch_servers" if shift == "lunch" else "hibachi_servers"
        reg_key = "lunch_servers" if shift == "lunch" else "dinner_servers"

        b_list = self.current_schedule[day].setdefault(b_key, [])

        if name in b_list:
            b_list.remove(name)
        else:
            if len(b_list) >= 3:
                messagebox.showinfo("B-side full",
                    "Maximum 3 B-side (sushi side) staff per shift.")
                return
            # Pull out of hibachi if present (B-side and hibachi are exclusive),
            # then backfill the vacated hibachi slot from the regular server pool.
            hib_list = self.current_schedule[day][hib_key]
            vacated_hibachi = name in hib_list
            if vacated_hibachi:
                hib_list.remove(name)
                if name not in self.current_schedule[day][reg_key]:
                    self.current_schedule[day][reg_key].append(name)
            b_list.append(name)
            if vacated_hibachi:
                self._backfill_hibachi(day, shift)

        # Re-run sidework letter assignment so W/X/Y reflect the change
        self.generator.schedule = self.current_schedule
        self.generator._assign_sidework_letters()
        self._display_schedule()

    def _backfill_hibachi(self, day, shift):
        """Promote regular servers into hibachi until the day's target is met.

        Selection priority:
        1. Servers below their personal weekly hibachi target (target=2 for
           most, target=1 for senior staff like Andy/Dian/Leony) — they need
           to fill their requirement first.
        2. Within that group, fewest current hibachi shifts first.
        3. Random tiebreaker for full ties.
        4. As a last resort, servers already at/over their target are still
           pickable so the day's hibachi requirement always gets filled. Within
           that fallback group, the LEAST senior staff get targeted first
           (so Andy at seniority 6 gets the extra shift before Leony/Dian at 10),
           with random tiebreaker for equal seniorities.

        Skips emergency_only and always_hibachi staff, plus anyone already
        marked B-side for that shift.
        """
        config = self.data.config
        if shift == "lunch":
            slot_target = config["staffing"][day].get("hibachi_lunch", 0)
            hib_list = self.current_schedule[day]["hibachi_lunch_servers"]
            reg_list = self.current_schedule[day]["lunch_servers"]
            b_set = set(self.current_schedule[day].get("b_side_lunch", []))
        else:
            slot_target = config["staffing"][day].get("hibachi_dinner", 0)
            hib_list = self.current_schedule[day]["hibachi_servers"]
            reg_list = self.current_schedule[day]["dinner_servers"]
            b_set = set(self.current_schedule[day].get("b_side_dinner", []))

        needed = slot_target - len(hib_list)
        if needed <= 0:
            return

        # Tally hibachi shifts across the week so we balance the swap-in choice
        week_hib_counts = {}
        for d in DAYS:
            for n in self.current_schedule[d]["hibachi_lunch_servers"]:
                week_hib_counts[n] = week_hib_counts.get(n, 0) + 1
            for n in self.current_schedule[d]["hibachi_servers"]:
                week_hib_counts[n] = week_hib_counts.get(n, 0) + 1

        eligible = []
        for n in reg_list:
            if n in b_set:
                continue
            staff = self.data.get_staff_by_name(n)
            if not staff:
                continue
            flags = staff.get("flags", [])
            if "emergency_only" in flags or "always_hibachi" in flags:
                continue
            eligible.append(n)

        # Shuffle first so full ties resolve randomly (stable sort preserves order otherwise)
        random.shuffle(eligible)

        def seniority_of(n):
            return (self.data.get_staff_by_name(n) or {}).get("seniority", 0)

        # Sort: under-target group first (by fewest count), then over-target
        # group (by lowest seniority — least senior absorbs the extra shift).
        # Random pre-shuffle handles ties within each tier.
        eligible.sort(key=lambda n: (
            0 if week_hib_counts.get(n, 0) < hibachi_target(n) else 1,
            week_hib_counts.get(n, 0) if week_hib_counts.get(n, 0) < hibachi_target(n) else seniority_of(n),
        ))

        for picked in eligible[:needed]:
            reg_list.remove(picked)
            hib_list.append(picked)

    def _edit_cell(self, day, key, index):
        """Let manager swap a staff member in a cell via searchable dropdown."""
        if not self.current_schedule:
            return

        if key == "dinner_servers":
            hibachi = self.current_schedule[day]["hibachi_servers"]
            servers = self.current_schedule[day]["dinner_servers"]
            combined = hibachi + servers
            if index >= len(combined):
                return
            current_name = combined[index]
        else:
            lst = self.current_schedule[day][key]
            if index >= len(lst):
                return
            current_name = lst[index]

        # Get available staff for this slot
        available = self._get_available_for_slot(day, key)
        display_names = ["(remove)"] + [label for _, label in available]
        name_map = {"(remove)": "(remove)"}
        for real_name, label in available:
            name_map[label] = real_name

        dlg = tk.Toplevel(self.root)
        dlg.title(f"Edit {day} - {key.replace('_', ' ').title()}")
        dlg.geometry("350x150")
        dlg.transient(self.root)
        dlg.grab_set()

        ttk.Label(dlg, text=f"Current: {current_name}",
                 font=("Helvetica", 11, "bold")).pack(pady=(15, 5))
        ttk.Label(dlg, text="Type to search, then select:").pack()

        combo_var = tk.StringVar()
        combo = ttk.Combobox(dlg, textvariable=combo_var, values=display_names,
                             width=35, font=("Helvetica", 10))
        combo.pack(padx=15, pady=10)
        combo.focus_set()

        # Filter dropdown as user types
        def on_keyrelease(event):
            if event.keysym in ('Return', 'Tab', 'Escape', 'Up', 'Down', 'Left', 'Right'):
                return
            typed = combo_var.get().lower()
            if not typed:
                combo["values"] = display_names
                return
            filtered = [n for n in display_names if n.lower().startswith(typed)]
            if not filtered:
                filtered = [n for n in display_names if typed in n.lower()]
            combo["values"] = filtered if filtered else display_names

        combo.bind("<KeyRelease>", on_keyrelease)
        combo.bind("<FocusIn>", lambda e: combo.configure(values=display_names))

        def apply(event=None):
            selection = combo_var.get()
            if not selection:
                dlg.destroy()
                return
            new_name = name_map.get(selection, selection)

            if key == "dinner_servers":
                hibachi = self.current_schedule[day]["hibachi_servers"]
                servers = self.current_schedule[day]["dinner_servers"]
                if index < len(hibachi):
                    if new_name == "(remove)":
                        hibachi.pop(index)
                    else:
                        hibachi[index] = new_name
                else:
                    adj_idx = index - len(hibachi)
                    if new_name == "(remove)":
                        servers.pop(adj_idx)
                    else:
                        servers[adj_idx] = new_name
            else:
                lst = self.current_schedule[day][key]
                if new_name == "(remove)":
                    lst.pop(index)
                else:
                    lst[index] = new_name

            # Clean up B-side membership for the removed/replaced name
            if new_name != current_name:
                for b_key in ("b_side_lunch", "b_side_dinner"):
                    b_list = self.current_schedule[day].get(b_key, [])
                    if current_name in b_list:
                        b_list.remove(current_name)
                self.generator.schedule = self.current_schedule
                self.generator._assign_sidework_letters()

            self._display_schedule()
            dlg.destroy()

        combo.bind("<<ComboboxSelected>>", apply)
        ttk.Button(dlg, text="Apply", command=apply).pack(pady=5)

    def _select_midshifts(self):
        """Open dialog for manager to pick midshift servers"""
        if not self.current_schedule:
            messagebox.showinfo("Info", "Generate a schedule first.")
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("Select Midshift Servers")
        dlg.geometry("700x500")
        dlg.transient(self.root)
        dlg.grab_set()

        ttk.Label(dlg, text="Check servers for midshift (* mark)",
                 font=("Helvetica", 11, "bold")).pack(pady=10)

        canvas = tk.Canvas(dlg)
        scrollbar = ttk.Scrollbar(dlg, orient=tk.VERTICAL, command=canvas.yview)
        inner = ttk.Frame(canvas)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Headers
        ttk.Label(inner, text="Server", width=12, font=("Helvetica", 9, "bold")).grid(
            row=0, column=0, padx=5, pady=3)
        for di, day in enumerate(DAYS):
            ttk.Label(inner, text=day, width=8, font=("Helvetica", 9, "bold")).grid(
                row=0, column=di + 1, padx=3, pady=3)

        mid_vars = {}
        # Get all servers who appear in any shift
        all_names = set()
        for day in DAYS:
            for name in self.current_schedule[day]["lunch_servers"]:
                all_names.add(name)
            for name in self.current_schedule[day]["hibachi_lunch_servers"]:
                all_names.add(name)
            for name in self.current_schedule[day]["dinner_servers"]:
                all_names.add(name)
            for name in self.current_schedule[day]["hibachi_servers"]:
                all_names.add(name)

        # Dian and Leony have fixed midshifts
        existing_mids = {}
        for day in DAYS:
            existing_mids[day] = self.current_schedule[day].get("mid_servers", [])

        for ri, name in enumerate(sorted(all_names)):
            ttk.Label(inner, text=name, width=12).grid(row=ri + 1, column=0, padx=5, pady=2, sticky="w")
            mid_vars[name] = {}
            for di, day in enumerate(DAYS):
                # Check if this person works this day
                works = (name in self.current_schedule[day]["lunch_servers"] or
                         name in self.current_schedule[day]["hibachi_lunch_servers"] or
                         name in self.current_schedule[day]["dinner_servers"] or
                         name in self.current_schedule[day]["hibachi_servers"])
                is_mid = name in existing_mids.get(day, [])
                var = tk.BooleanVar(value=is_mid)
                mid_vars[name][day] = var
                cb = ttk.Checkbutton(inner, variable=var)
                if not works:
                    cb.configure(state="disabled")
                # Dian/Leony midshifts are fixed
                staff = self.data.get_staff_by_name(name)
                if staff and staff.get("fixed_schedule"):
                    cb.configure(state="disabled")
                cb.grid(row=ri + 1, column=di + 1, padx=3, pady=2)

        def apply_mids():
            for day in DAYS:
                mids = []
                for name, days in mid_vars.items():
                    if day in days and days[day].get():
                        mids.append(name)
                self.current_schedule[day]["mid_servers"] = mids
            self._display_schedule()
            dlg.destroy()

        ttk.Button(dlg, text="Apply Midshifts", command=apply_mids).pack(pady=10)

    # ---- TAB 5: EXPORT ----
    def _build_export_tab(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="  Export  ")

        ttk.Label(frame, text="Export Schedule",
                 font=("Helvetica", 14, "bold")).pack(pady=20)

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(pady=20)

        ttk.Button(btn_frame, text="Export to Excel (.xlsx)",
                  command=self._export_excel).pack(pady=10, ipadx=20, ipady=10)
        ttk.Button(btn_frame, text="Export to PDF",
                  command=self._export_pdf).pack(pady=10, ipadx=20, ipady=10)

        # Status
        self.export_status = tk.StringVar(value="Generate a schedule first, then export.")
        ttk.Label(frame, textvariable=self.export_status,
                 font=("Helvetica", 10)).pack(pady=20)

    def _get_week_start_date(self):
        date_str = self.week_start_var.get().strip()
        if date_str:
            try:
                return datetime.strptime(date_str, "%Y-%m-%d")
            except ValueError:
                pass
        return None

    def _export_excel(self):
        if not self.current_schedule:
            messagebox.showinfo("Info", "Generate a schedule first.")
            return

        week_date = self._get_week_start_date()
        default_name = "Schedule"
        if week_date:
            end_date = week_date + timedelta(days=6)
            default_name = f"{week_date.strftime('%Y %b %d')} - {end_date.strftime('%d')}"

        filepath = filedialog.asksaveasfilename(
            title="Save Excel Schedule",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialdir=self.data.config.get("export_path", "") or str(APP_DIR),
            initialfile=default_name,
        )
        if not filepath:
            return

        # Save the export path for next time
        self.data.config["export_path"] = str(Path(filepath).parent)
        self.data.save_config()

        try:
            exporter = ExcelExporter(self.current_schedule, self.data.config, week_date)
            exporter.export(filepath)
            self.export_status.set(f"Excel saved: {filepath}")
            messagebox.showinfo("Success", f"Schedule exported to:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Error", f"Export failed:\n{e}")

    def _export_pdf(self):
        if not self.current_schedule:
            messagebox.showinfo("Info", "Generate a schedule first.")
            return

        if FPDF is None:
            messagebox.showerror("Error", "fpdf2 not installed. Run setup.sh to install.")
            return

        week_date = self._get_week_start_date()
        default_name = "Schedule"
        if week_date:
            end_date = week_date + timedelta(days=6)
            default_name = f"{week_date.strftime('%Y %b %d')} - {end_date.strftime('%d')}"

        filepath = filedialog.asksaveasfilename(
            title="Save PDF Schedule",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialdir=self.data.config.get("export_path", "") or str(APP_DIR),
            initialfile=default_name,
        )
        if not filepath:
            return

        self.data.config["export_path"] = str(Path(filepath).parent)
        self.data.save_config()

        try:
            exporter = PDFExporter(self.current_schedule, self.data.config, week_date)
            exporter.export(filepath)
            self.export_status.set(f"PDF saved: {filepath}")
            messagebox.showinfo("Success", f"Schedule exported to:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Error", f"Export failed:\n{e}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def main():
    root = tk.Tk()
    app = ToyoSchedulerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
